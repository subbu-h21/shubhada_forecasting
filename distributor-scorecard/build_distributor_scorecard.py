r"""
Distributor (Supplier) Profit Scorecard
=========================================
Standalone, READ-ONLY analysis of the pharmacy-reckoner PURCHASE data. For
each distributor, computes the "potential"/embedded profit locked into what
we bought from them: if every unit received (including free-scheme goods)
were sold at MRP, what margin does that represent versus what we actually
paid.

This is a BUYING/negotiation KPI - which distributors give the best terms -
NOT realized P&L (most of the stock this measures is still unsold).

Does not import or modify anything in the pharmacy-reckoner project. It only
reads the CSV path below; nothing is written back to it.

Usage:
    python build_distributor_scorecard.py

Output:
    Distributor Scorecard.xlsx (written next to this script)
"""
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PURCHASE_CSV = Path(r'D:\claude projects\project2\pharmacy-reckoner\data\processed\purchase_master.csv')
OUTPUT_PATH = Path(__file__).parent / 'Distributor Scorecard.xlsx'
TOP_N_PRODUCTS = 15


# ---------------------------------------------------------------------------
# Load + per-line embedded-profit calculation
# ---------------------------------------------------------------------------
def load_purchases():
    # Inv.No mixes plain ints, leading-zero strings, and slash-suffixed
    # formats ('1274/26-27') across suppliers - read as string throughout so
    # nothing gets silently coerced or mismatched.
    return pd.read_csv(PURCHASE_CSV, dtype={'Inv.No': str})


def safe_pct(numerator, denominator):
    denominator = np.asarray(denominator, dtype=float)
    numerator = np.asarray(numerator, dtype=float)
    return np.where(denominator > 0, 100 * numerator / np.where(denominator == 0, 1, denominator), 0.0)


def compute_lines(df):
    """Per purchase-line embedded profit, pre-tax both sides. Lines with
    MRP <= 0 can't be valued (no reference sell price) and are excluded -
    returned separately so the caller can report what was left out."""
    df = df.copy()
    df['Free Qty'] = df['Free Qty'].fillna(0)
    df['Disc Amount'] = df['Disc Amount'].fillna(0)

    excluded = df[df['MRP'] <= 0].copy()
    valid = df[df['MRP'] > 0].copy()

    valid['Units_Received'] = valid['Qty'] + valid['Free Qty']
    valid['Max_Sell_Pretax'] = (valid['MRP'] / (1 + valid['Tax Rate'] / 100)) * valid['Units_Received']
    valid['Net_Cost_Pretax'] = valid['Qty'] * valid['Sale Rate'] - valid['Disc Amount']
    valid['Embedded_Profit'] = valid['Max_Sell_Pretax'] - valid['Net_Cost_Pretax']

    return valid, excluded


# ---------------------------------------------------------------------------
# Roll-ups: line -> invoice -> supplier x month -> supplier (all-history)
# ---------------------------------------------------------------------------
def build_distributor_summary(lines):
    g = lines.groupby('Supplier').agg(
        Invoices=('Inv.No', 'nunique'),
        Lines=('Product', 'count'),
        Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'),
        Embedded_Profit=('Embedded_Profit', 'sum'),
        Months_Active=('Source_Month', 'nunique'),
    ).reset_index()
    # "Total invoice value" = actual tax-inclusive spend (sum of Item Total),
    # not the repeated per-invoice Invoice Amount column.
    inv_value = lines.groupby('Supplier')['Item Total'].sum().rename('Total_Invoice_Value')
    g = g.merge(inv_value, on='Supplier')
    g['Margin_Pct'] = safe_pct(g['Embedded_Profit'], g['Max_Sell_Pretax'])
    return g.sort_values('Embedded_Profit', ascending=False).reset_index(drop=True)


def build_distributor_month(lines):
    g = lines.groupby(['Supplier', 'Source_Month']).agg(
        Invoice_Value=('Item Total', 'sum'),
        Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'),
        Embedded_Profit=('Embedded_Profit', 'sum'),
    ).reset_index()
    g['Margin_Pct'] = safe_pct(g['Embedded_Profit'], g['Max_Sell_Pretax'])
    supplier_order = lines.groupby('Supplier')['Embedded_Profit'].sum().sort_values(ascending=False).index
    g['Supplier'] = pd.Categorical(g['Supplier'], categories=supplier_order, ordered=True)
    return g.sort_values(['Supplier', 'Source_Month']).reset_index(drop=True), list(supplier_order)


def build_per_invoice(lines):
    g = lines.groupby(['Inv.No', 'Supplier']).agg(
        Date=('Date', 'first'),
        Source_Month=('Source_Month', 'first'),
        Lines=('Product', 'count'),
        Invoice_Value=('Item Total', 'sum'),
        Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'),
        Embedded_Profit=('Embedded_Profit', 'sum'),
    ).reset_index()
    g['Margin_Pct'] = safe_pct(g['Embedded_Profit'], g['Max_Sell_Pretax'])
    return g.sort_values('Embedded_Profit', ascending=False).reset_index(drop=True)


def build_top_products(lines, supplier_order, top_n):
    g = lines.groupby(['Supplier', 'Product']).agg(
        Units_Received=('Units_Received', 'sum'),
        Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'),
        Embedded_Profit=('Embedded_Profit', 'sum'),
    ).reset_index()
    g['Margin_Pct'] = safe_pct(g['Embedded_Profit'], g['Max_Sell_Pretax'])

    parts = []
    for supplier in supplier_order:
        sub = g[g['Supplier'] == supplier].sort_values('Embedded_Profit', ascending=False).head(top_n)
        parts.append(sub)
    return pd.concat(parts, ignore_index=True) if parts else g.iloc[0:0]


def build_entered_by(lines):
    e = lines[lines['Entered By'].notna()].copy()
    if e.empty or len(e) < 0.1 * len(lines):
        return None  # column effectively unused - skip the bonus sheet
    e['Entered By'] = e['Entered By'].str.strip().str.title()
    g = e.groupby('Entered By').agg(
        Lines=('Product', 'count'),
        Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'),
        Embedded_Profit=('Embedded_Profit', 'sum'),
    ).reset_index()
    g['Margin_Pct'] = safe_pct(g['Embedded_Profit'], g['Max_Sell_Pretax'])
    return g.sort_values('Embedded_Profit', ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Validation - reproduce the two known reference figures before trusting the
# rest of the output.
# ---------------------------------------------------------------------------
def run_validation(lines):
    print('\n' + '=' * 72)
    print('VALIDATION')
    print('=' * 72)

    inv = lines[(lines['Inv.No'] == '1839') & (lines['Supplier'] == 'TULASI PHARMA')]
    ms, nc = inv['Max_Sell_Pretax'].sum(), inv['Net_Cost_Pretax'].sum()
    ep = ms - nc
    margin = 100 * ep / ms if ms else 0
    print(f"\nInvoice 1839 (TULASI PHARMA):")
    print(f"  {'Metric':<20}{'Expected':>16}{'Computed':>16}")
    print(f"  {'Max sell (pre-tax)':<20}{10914.71:>16,.2f}{ms:>16,.2f}")
    print(f"  {'Net cost (pre-tax)':<20}{8592.54:>16,.2f}{nc:>16,.2f}")
    print(f"  {'Embedded profit':<20}{2322.17:>16,.2f}{ep:>16,.2f}")
    print(f"  {'Margin %':<20}{21.3:>15.1f}%{margin:>15.1f}%")

    kap = lines[lines['Supplier'] == 'KAPILA MEDICAL AGENCIES']
    kap_ms, kap_nc = kap['Max_Sell_Pretax'].sum(), kap['Net_Cost_Pretax'].sum()
    kap_ep = kap_ms - kap_nc
    kap_margin = 100 * kap_ep / kap_ms if kap_ms else 0
    aug = kap[kap['Source_Month'] == '2026-08']
    aug_inv_value = aug['Item Total'].sum()
    aug_ep = aug['Max_Sell_Pretax'].sum() - aug['Net_Cost_Pretax'].sum()
    print(f"\nKAPILA MEDICAL AGENCIES (top distributor by embedded profit, all-history):")
    print(f"  {'Metric':<28}{'Expected':>16}{'Computed':>16}")
    print(f"  {'All-history margin %':<28}{'~28%':>16}{kap_margin:>15.1f}%")
    print(f"  {'2026-08 invoice value':<28}{2603141:>16,.0f}{aug_inv_value:>16,.2f}")
    print(f"  {'2026-08 embedded profit':<28}{978936:>16,.0f}{aug_ep:>16,.2f}")
    print('=' * 72)


# ---------------------------------------------------------------------------
# Excel report builder
# ---------------------------------------------------------------------------
FONT = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=11)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color='555555')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_table_counter = [0]


def write_df(ws, df, start_row=1, money_cols=None, qty_cols=None):
    money_cols = money_cols or []
    qty_cols = qty_cols or []
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            colname = df.columns[j - 1]
            if colname in money_cols:
                c.number_format = '#,##0.00'
            if colname in qty_cols:
                c.number_format = '#,##0'
    last_row = start_row + len(df)
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        _table_counter[0] += 1
        tab = Table(displayName=f'T{_table_counter[0]}', ref=f'A{start_row}:{last_col}{last_row}')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
        ws.add_table(tab)
    return last_row


def autosize(ws, widths):
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build_report(lines, excluded, dist_summary, dist_month, supplier_order, per_invoice, top_products, entered_by, months):
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    ws['B2'] = 'Distributor (Supplier) Profit Scorecard'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = f'Purchase history: {", ".join(months)}'
    ws['B3'].font = SUBTITLE_FONT

    notes = [
        '',
        'WHAT THIS MEASURES',
        '- "Embedded profit" is the POTENTIAL margin locked into what we bought from each distributor:',
        '  if every unit received (including free-scheme goods) were sold at MRP, what would that be',
        '  worth versus what we actually paid.',
        '- This is a BUYING / negotiation KPI - which distributors give the best terms - NOT realized',
        '  P&L. Most of this stock has not actually been sold yet.',
        '',
        'RULES USED',
        '- Pre-tax on both sides: MRP (tax-inclusive) has GST stripped out before comparing to cost.',
        '- Bill Disc (invoice-level discount) is ignored entirely - only line-level PTR and Disc Amount count.',
        '- Free-scheme goods (Free Qty) are included as received units at zero extra cost.',
        '- Lines with MRP <= 0 cannot be valued and are excluded (see count below).',
        '',
        f'Suppliers analyzed: {dist_summary["Supplier"].nunique()}',
        f'Purchase lines used: {len(lines):,}',
        f'Lines excluded (MRP missing/zero): {len(excluded):,}  '
        f'(Item Total of excluded lines: Rs {excluded["Item Total"].sum():,.2f})',
        f'"Top Products per Distributor" sheet shows up to {TOP_N_PRODUCTS} products per distributor.',
    ]
    r = 5
    for line in notes:
        bold = line.isupper() and line != ''
        ws.cell(row=r, column=2, value=line).font = Font(name=FONT, bold=bold, size=10)
        r += 1

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 100

    # ---- Distributor Summary ----
    ws = wb.create_sheet('Distributor Summary')
    ws.sheet_view.showGridLines = False
    df1 = dist_summary.rename(columns={
        'Total_Invoice_Value': 'Total Invoice Value', 'Max_Sell_Pretax': 'Max Sell Value (pre-tax)',
        'Net_Cost_Pretax': 'Net Cost (pre-tax)', 'Embedded_Profit': 'Embedded Profit',
        'Margin_Pct': 'Margin %', 'Months_Active': 'Months Active'})
    df1 = df1[['Supplier', 'Invoices', 'Lines', 'Total Invoice Value', 'Max Sell Value (pre-tax)',
               'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %', 'Months Active']]
    write_df(ws, df1, start_row=1,
             money_cols=['Total Invoice Value', 'Max Sell Value (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
             qty_cols=['Invoices', 'Lines', 'Months Active'])
    autosize(ws, [34, 11, 10, 18, 20, 18, 16, 10, 12])
    ws.freeze_panes = 'A2'

    # ---- Distributor x Month ----
    ws = wb.create_sheet('Distributor x Month')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Supplier x Month detail'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    df2 = dist_month.rename(columns={'Source_Month': 'Month', 'Invoice_Value': 'Invoice Value',
                                      'Max_Sell_Pretax': 'Max Sell (pre-tax)', 'Net_Cost_Pretax': 'Net Cost (pre-tax)',
                                      'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
    df2 = df2[['Supplier', 'Month', 'Invoice Value', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
    df2['Supplier'] = df2['Supplier'].astype(str)
    last2 = write_df(ws, df2, start_row=2,
                      money_cols=['Invoice Value', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'])

    r3 = last2 + 3
    ws.cell(row=r3, column=1, value='Pivot: Embedded Profit by Supplier x Month (suppliers ordered by all-history total)').font = \
        Font(name=FONT, bold=True, size=11)
    pivot = dist_month.pivot(index='Supplier', columns='Source_Month', values='Embedded_Profit').fillna(0)
    pivot = pivot.reindex(supplier_order)
    pivot['Total'] = pivot.sum(axis=1)
    pivot = pivot.reset_index()
    pivot['Supplier'] = pivot['Supplier'].astype(str)
    write_df(ws, pivot, start_row=r3 + 1, money_cols=list(pivot.columns[1:]))
    autosize(ws, [34, 16, 16, 16, 16, 16])
    ws.freeze_panes = 'A3'

    # ---- Per Invoice ----
    ws = wb.create_sheet('Per Invoice')
    ws.sheet_view.showGridLines = False
    df3 = per_invoice.rename(columns={
        'Inv.No': 'Inv.No', 'Source_Month': 'Month', 'Invoice_Value': 'Invoice Value',
        'Max_Sell_Pretax': 'Max Sell (pre-tax)', 'Net_Cost_Pretax': 'Net Cost (pre-tax)',
        'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
    df3 = df3[['Date', 'Inv.No', 'Supplier', 'Month', 'Lines', 'Invoice Value',
               'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
    write_df(ws, df3, start_row=1,
             money_cols=['Invoice Value', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
             qty_cols=['Lines'])
    autosize(ws, [12, 16, 30, 10, 8, 16, 16, 16, 16, 10])
    ws.freeze_panes = 'A2'

    # ---- Top Products per Distributor ----
    ws = wb.create_sheet('Top Products per Distributor')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f'Top {TOP_N_PRODUCTS} products by embedded profit, per distributor (suppliers ordered by all-history total)'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    df4 = top_products.rename(columns={
        'Units_Received': 'Units Received', 'Max_Sell_Pretax': 'Max Sell (pre-tax)',
        'Net_Cost_Pretax': 'Net Cost (pre-tax)', 'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
    df4 = df4[['Supplier', 'Product', 'Units Received', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
    write_df(ws, df4, start_row=2,
             money_cols=['Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
             qty_cols=['Units Received'])
    autosize(ws, [30, 32, 14, 16, 16, 16, 10])
    ws.freeze_panes = 'A3'

    # ---- Entered By (optional bonus) ----
    if entered_by is not None:
        ws = wb.create_sheet('Entered By')
        ws.sheet_view.showGridLines = False
        ws['A1'] = 'Embedded profit / margin of purchase entries, by who keyed them in (name normalized: trimmed + title case)'
        ws['A1'].font = Font(name=FONT, bold=True, size=11)
        df5 = entered_by.rename(columns={'Max_Sell_Pretax': 'Max Sell (pre-tax)', 'Net_Cost_Pretax': 'Net Cost (pre-tax)',
                                          'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
        df5 = df5[['Entered By', 'Lines', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
        write_df(ws, df5, start_row=2,
                 money_cols=['Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
                 qty_cols=['Lines'])
        autosize(ws, [26, 10, 16, 16, 16, 10])
        ws.freeze_panes = 'A3'

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


def main():
    df = load_purchases()
    lines, excluded = compute_lines(df)
    months = sorted(lines['Source_Month'].unique())

    print(f'Months covered: {months}')
    print(f'Suppliers: {lines["Supplier"].nunique()}')
    print(f'Lines used: {len(lines):,}  |  Lines excluded (MRP<=0): {len(excluded):,} '
          f'(Item Total Rs {excluded["Item Total"].sum():,.2f})')

    dist_summary = build_distributor_summary(lines)
    dist_month, supplier_order = build_distributor_month(lines)
    per_invoice = build_per_invoice(lines)
    top_products = build_top_products(lines, supplier_order, TOP_N_PRODUCTS)
    entered_by = build_entered_by(lines)

    run_validation(lines)

    build_report(lines, excluded, dist_summary, dist_month, supplier_order, per_invoice, top_products, entered_by, months)
    print(f'\nReport saved: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
