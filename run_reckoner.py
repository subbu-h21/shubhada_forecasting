r"""
Pharmacy Ready Reckoner
========================
Run this every month after dropping the new Sale and Purchase XLS files into
data\raw\. It remembers every month you've ever fed it (stored in
data\processed\) and regenerates all reports using the FULL history each time.

Usage:
    python run_reckoner.py

Output:
    reports\<latest-month>\Monthly Reckoner Report.xlsx
"""
import hashlib
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference

ROOT = Path(__file__).parent
RAW_DIR = ROOT / 'data' / 'raw'
PROC_DIR = ROOT / 'data' / 'processed'
REPORTS_DIR = ROOT / 'reports'
MANIFEST_PATH = PROC_DIR / 'manifest.json'
SALES_MASTER = PROC_DIR / 'sales_master.csv'
PURCH_MASTER = PROC_DIR / 'purchase_master.csv'
HOLIDAYS_PATH = ROOT / 'holidays.json'


def load_holidays():
    """Declared holidays as {date_string: name}. Edit holidays.json (or mark
    them from the mobile calendar) to add festival/regional dates - lunar-
    calendar festivals (Diwali, Ganesh Chaturthi, Eid, etc.) shift every
    year, so they're deliberately not guessed here; add the confirmed date
    for the specific year yourself."""
    if not HOLIDAYS_PATH.exists():
        return {}
    entries = json.loads(HOLIDAYS_PATH.read_text())
    return {e['date']: e['name'] for e in entries}


def save_holiday(date_str, name):
    """Add or update one holiday. Pass name=None to remove that date."""
    entries = []
    if HOLIDAYS_PATH.exists():
        entries = json.loads(HOLIDAYS_PATH.read_text())
    entries = [e for e in entries if e['date'] != date_str]
    if name:
        entries.append({'date': date_str, 'name': name})
    entries.sort(key=lambda e: e['date'])
    HOLIDAYS_PATH.write_text(json.dumps(entries, indent=2))


def is_nth_weekday_of_month(d, weekday, n):
    """True if date d is the nth occurrence of `weekday` (0=Mon..6=Sun) in its month."""
    return d.weekday() == weekday and (d.day - 1) // 7 == n - 1


def classify_special_day(d, holidays):
    """Returns (day_type, label) for a date. day_type is one of:
    'holiday', 'pre_holiday', 'post_holiday', 'second_saturday', or None.
    Priority: an actual declared holiday wins over being adjacent to one;
    adjacency wins over the (data-driven, not assumed) 2nd-Saturday effect."""
    d = pd.Timestamp(d)  # normalize: a bare date object or a Timestamp with a
    # time component both need to collapse to plain 'YYYY-MM-DD' for the
    # holidays-dict lookup below - str() on a raw Timestamp keeps "00:00:00"
    # and would silently never match.
    ds = str(d.date())
    if ds in holidays:
        return 'holiday', holidays[ds]
    nxt, prv = str((d + pd.Timedelta(days=1)).date()), str((d - pd.Timedelta(days=1)).date())
    if nxt in holidays:
        return 'pre_holiday', f'Day before {holidays[nxt]}'
    if prv in holidays:
        return 'post_holiday', f'Day after {holidays[prv]}'
    if is_nth_weekday_of_month(d, 5, 2):  # 5 = Saturday
        return 'second_saturday', '2nd Saturday'
    return None, ''

SALE_SIGNATURE = {'Patient', 'Product', 'Qty', 'Item Total'}
PURCH_SIGNATURE = {'Supplier', 'Invoice Amount', 'Product', 'Qty'}

# Sales Inv.No carries a branch code prefix (e.g. '2627WH34042' -> '2627WH').
# Purchase invoices do NOT carry this - purchases can't be split by branch.
# The numeric part ('2627') looks like a year/series code that will change
# next financial year; only the trailing letter reliably identifies the
# branch. Add new codes here as they show up - unmapped codes are labelled
# 'Unknown branch (<code>)' rather than dropped, so nothing goes missing.
BRANCH_MAP = {
    '2627WS': 'Shivaji Chowk',
    '2627WH': 'Hospet Road',
    '2627WB': 'Herur',
}
BRANCH_PREFIX_RE = re.compile(r'^(\d+[A-Za-z]+)')


def extract_branch(inv_no):
    m = BRANCH_PREFIX_RE.match(str(inv_no))
    if not m:
        return 'Unknown branch'
    code = m.group(1).upper()
    return BRANCH_MAP.get(code, f'Unknown branch ({code})')


# Columns not present in any export yet (as of building this) - employee and
# customer analysis activates automatically once a future export includes
# one of these under any of its common naming variants. Nothing breaks in
# the meantime; ingestion already keeps whatever columns a file has.
MOBILE_CANDIDATES = ['Mobile Number', 'Mobile No', 'Mobile', 'Phone Number', 'Phone', 'Patient Mobile', 'Contact Number']
TIMESTAMP_CANDIDATES = ['Time Stamp', 'Timestamp', 'Time', 'Bill Time']
BILLED_BY_CANDIDATES = ['Billed By', 'Billed by', 'Biller', 'Cashier', 'Bill By']
GIVEN_BY_CANDIDATES = ['Item Given By', 'Given By', 'Given by', 'Dispensed By', 'Issued By']
CREATED_BY_CANDIDATES = ['Created By', 'Created by', 'Entry By', 'Entered By']


def find_col(df, candidates):
    """First matching column name (from a list of naming variants) that
    actually has real data in it - a present-but-empty column doesn't count."""
    for c in candidates:
        if c in df.columns and df[c].notna().any():
            return c
    return None


# Individual-unit quantities (tablets/ml/etc.) aren't how a pharmacist
# actually thinks about stock - strips are. Qty / Factor converts to strips
# wherever a product-level quantity is displayed. Factor is a physical
# packaging attribute of the product, so one representative value (the most
# common one seen) is used regardless of which month/invoice it came from.
def build_product_factor_map(sales, purch=None):
    f1 = sales.groupby('Product')['Factor'].agg(lambda s: s.mode().iat[0] if len(s.mode()) else 1)
    if purch is not None:
        f2 = purch.groupby('Product')['Factor'].agg(lambda s: s.mode().iat[0] if len(s.mode()) else 1)
        f1 = f1.combine_first(f2)
    return f1.replace(0, 1).fillna(1).to_dict()


def to_strips(qty_series, product_series, factor_map):
    factors = product_series.map(factor_map).fillna(1).replace(0, 1)
    return (qty_series / factors).round(1)


MONTH_DAYS = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}


def days_in_month(year, month):
    if month == 2 and (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)):
        return 29
    return MONTH_DAYS[month]


def file_hash(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        h.update(f.read())
    return h.hexdigest()


def load_manifest():
    if MANIFEST_PATH.exists():
        m = json.loads(MANIFEST_PATH.read_text())
        m.setdefault('files', {})
        m.setdefault('chosen', {})
        return m
    return {'files': {}, 'chosen': {}}


def save_manifest(m):
    PROC_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(json.dumps(m, indent=2))


def fix_date_column(df, dominant_year, dominant_month):
    """Sales/purchase registers occasionally have a stray row keyed in
    MM/DD/YYYY instead of DD/MM/YYYY. Parse normally, then for any row whose
    month doesn't match the file's dominant month, try swapping day/month and
    keep the swap only if it lands in the dominant month."""
    parsed = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
    bad_idx = parsed.index[parsed.dt.month != dominant_month]
    if len(bad_idx) > 0:
        alt = pd.to_datetime(df.loc[bad_idx, 'Date'], dayfirst=False, errors='coerce')
        use_alt = alt.index[(alt.dt.month == dominant_month) & (alt.dt.year == dominant_year)]
        parsed.loc[use_alt] = alt.loc[use_alt]
    return parsed


def detect_type_and_load(path):
    df = pd.read_excel(path)
    cols = set(df.columns)
    if SALE_SIGNATURE.issubset(cols):
        kind = 'sale'
    elif PURCH_SIGNATURE.issubset(cols):
        kind = 'purchase'
    else:
        raise ValueError(f'{path.name}: could not recognize columns as Sale or Purchase register')

    raw_parsed = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
    ym_counts = raw_parsed.dt.to_period('M').value_counts()
    dominant_period = ym_counts.index[0]
    fixed_dates = fix_date_column(df, dominant_period.year, dominant_period.month)
    df['Date'] = fixed_dates
    df['Source_Month'] = str(dominant_period)
    df['Source_File'] = path.name

    in_month = df[(df['Date'].dt.year == dominant_period.year) & (df['Date'].dt.month == dominant_period.month)]
    unique_days = int(in_month['Date'].dt.date.nunique())
    coverage = {'unique_days': unique_days, 'rows': len(df)}
    return kind, df, coverage


def _remove_file_rows(master_path, filename):
    if not master_path.exists():
        return
    existing = pd.read_csv(master_path)
    if 'Source_File' in existing.columns and (existing['Source_File'] == filename).any():
        existing = existing[existing['Source_File'] != filename]
        existing.to_csv(master_path, index=False)


def _append_rows(master_path, df):
    if master_path.exists():
        existing = pd.read_csv(master_path)
        existing = existing[existing['Source_File'] != df['Source_File'].iloc[0]]
        combined = pd.concat([existing, df], ignore_index=True)
    else:
        combined = df
    combined.to_csv(master_path, index=False)


def ingest():
    manifest = load_manifest()
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    PROC_DIR.mkdir(parents=True, exist_ok=True)
    master_for = {'sale': SALES_MASTER, 'purchase': PURCH_MASTER}
    changed = False

    for path in sorted(RAW_DIR.glob('*.xlsx')):
        if path.name.startswith('~$'):
            continue
        h = file_hash(path)
        prior_meta = manifest['files'].get(path.name)
        if prior_meta and prior_meta.get('hash') == h:
            continue  # this exact file already processed, nothing changed

        kind, df, coverage = detect_type_and_load(path)
        month = df['Source_Month'].iloc[0]
        key = f'{kind}:{month}'
        chosen_name = manifest['chosen'].get(key)

        if chosen_name is None or chosen_name == path.name:
            # first file seen for this month+type, or this exact file was updated in place
            replace, reason = True, 'first file for this month' if chosen_name is None else 'file updated in place'
        else:
            prior_cov = manifest['files'].get(chosen_name, {})
            prior_days = prior_cov.get('unique_days', 0)
            prior_rows = prior_cov.get('rows', 0)
            if (coverage['unique_days'], coverage['rows']) > (prior_days, prior_rows):
                replace = True
                reason = f"wider date coverage ({coverage['unique_days']} days) than '{chosen_name}' ({prior_days} days) - replacing it"
            else:
                replace = False
                reason = f"narrower/equal date coverage ({coverage['unique_days']} days) vs current '{chosen_name}' ({prior_days} days) - skipped as duplicate month"

        if replace:
            if chosen_name and chosen_name != path.name:
                _remove_file_rows(master_for[kind], chosen_name)
                if chosen_name in manifest['files']:
                    manifest['files'][chosen_name]['status'] = 'superseded'
            _append_rows(master_for[kind], df)
            manifest['chosen'][key] = path.name
            manifest['files'][path.name] = {'hash': h, 'kind': kind, 'month': month, 'status': 'active', **coverage}
            print(f"Ingested {path.name} as {kind.upper()} -> {month} ({coverage['rows']} rows, {coverage['unique_days']} days): {reason}")
        else:
            manifest['files'][path.name] = {'hash': h, 'kind': kind, 'month': month, 'status': 'rejected', **coverage}
            print(f"Skipped {path.name}: {reason}")
        changed = True

    save_manifest(manifest)
    if not changed:
        print('No new files found in data\\raw - nothing to ingest.')

    sales = pd.read_csv(SALES_MASTER) if SALES_MASTER.exists() else pd.DataFrame()
    purch = pd.read_csv(PURCH_MASTER) if PURCH_MASTER.exists() else pd.DataFrame()
    return sales, purch


def next_month_str(ym_str):
    y, m = map(int, ym_str.split('-'))
    if m == 12:
        return f'{y+1}-01'
    return f'{y}-{m+1:02d}'


# ---------------------------------------------------------------------------
# Analysis 1: Demand forecast for the month AFTER the latest data
# ---------------------------------------------------------------------------
def build_demand_forecast(sales):
    months = sorted(sales['Source_Month'].unique())
    latest = months[-1]
    target_month = next_month_str(latest)
    ty, tm = map(int, target_month.split('-'))
    target_days = days_in_month(ty, tm)

    monthly = sales.groupby(['Product', 'Source_Month']).agg(
        Qty=('Qty', 'sum'), Value=('Item Total', 'sum')).reset_index()

    # Use the number of days actually COVERED by data for each month, not the
    # calendar length of the month. If the latest month was uploaded before
    # it finished (e.g. dropped in on the 25th), dividing by the full 30/31
    # understates its true daily rate and throws off the growth trend.
    dates = sales.copy()
    dates['Date'] = pd.to_datetime(dates['Date'], format='mixed')
    days_covered = dates.groupby('Source_Month')['Date'].apply(lambda d: d.dt.date.nunique())

    per_day = {}
    for m in months:
        d = max(1, days_covered.get(m, 1))
        sub = monthly[monthly['Source_Month'] == m].set_index('Product')
        per_day[m] = (sub['Qty'] / d)

    products = monthly['Product'].unique()
    rows = []
    for prod in products:
        series = [per_day[m].get(prod, 0.0) for m in months]
        total_qty = monthly[monthly['Product'] == prod]['Qty'].sum()
        n = len(series)

        if n == 1:
            flag = 'Only 1 month history'
            pred_per_day = series[0]
        elif series[-2] == 0 and series[-1] > 0:
            flag = 'New (no sales until latest month)'
            pred_per_day = series[-1]
        elif series[-2] > 0 and series[-1] == 0:
            flag = 'Stopped in latest month (check stock)'
            pred_per_day = 0.0
        elif total_qty < 10:
            flag = 'Low volume'
            pred_per_day = float(np.mean(series))
        else:
            growth = (series[-1] - series[-2]) / series[-2] if series[-2] > 0 else 0
            growth = max(-0.3, min(0.5, growth))
            pred_per_day = series[-1] * (1 + growth)
            flag = 'Growing' if growth > 0.05 else ('Declining' if growth < -0.05 else 'Stable')

        pred_qty = max(0, round(pred_per_day * target_days))
        rows.append((prod, flag, pred_qty))

    df = pd.DataFrame(rows, columns=['Product', 'Trend', 'Predicted_Qty'])

    price = sales.sort_values('Source_Month').groupby('Product').apply(
        lambda d: (d['Item Total'].sum() / d['Qty'].sum()) if d['Qty'].sum() > 0 else 0,
        include_groups=False)
    df['Avg_Price'] = df['Product'].map(price).fillna(0)
    df['Predicted_Value'] = (df['Predicted_Qty'] * df['Avg_Price']).round(2)
    factor_map = build_product_factor_map(sales)
    df['Predicted_Qty_Strips'] = to_strips(df['Predicted_Qty'], df['Product'], factor_map)
    df = df.sort_values('Predicted_Value', ascending=False).reset_index(drop=True)
    return df, target_month, months


# ---------------------------------------------------------------------------
# Analysis 1b: Branch-wise summary and per-branch demand forecast
#
# Only possible on the Sales side - Purchase invoices carry no branch code,
# so Over-Under Purchased and Profit & Margin cannot be split by branch.
# ---------------------------------------------------------------------------
def build_branch_report(sales):
    sales = sales.copy()
    sales['Branch'] = sales['Inv.No'].apply(extract_branch)
    months = sorted(sales['Source_Month'].unique())

    monthly = sales.groupby(['Branch', 'Source_Month']).agg(
        Revenue=('Item Total', 'sum'), Invoices=('Inv.No', 'nunique'), Patients=('Patient', 'nunique')).reset_index()
    pivot = monthly.pivot(index='Branch', columns='Source_Month', values='Revenue').fillna(0)
    pivot.columns = [f'Revenue_{m}' for m in pivot.columns]

    totals = sales.groupby('Branch').agg(
        Total_Revenue=('Item Total', 'sum'), Invoices=('Inv.No', 'nunique'), Patients=('Patient', 'nunique')).reset_index()
    summary = totals.merge(pivot, on='Branch', how='left')

    if len(months) >= 2:
        first_col, last_col = f'Revenue_{months[0]}', f'Revenue_{months[-1]}'
        summary['Growth_Pct'] = np.where(summary[first_col] > 0,
                                          ((summary[last_col] - summary[first_col]) / summary[first_col] * 100).round(1), 0)
    else:
        summary['Growth_Pct'] = 0
    summary = summary.sort_values('Total_Revenue', ascending=False).reset_index(drop=True)

    per_branch_forecast = []
    for branch in summary['Branch']:
        sub = sales[sales['Branch'] == branch].drop(columns=['Branch'])
        fc, target_month, _ = build_demand_forecast(sub)
        fc.insert(0, 'Branch', branch)
        per_branch_forecast.append(fc)
    branch_forecast = pd.concat(per_branch_forecast, ignore_index=True) if per_branch_forecast else pd.DataFrame()

    return summary, branch_forecast


# ---------------------------------------------------------------------------
# Analysis 1c: Daily footfall (unique bills per day) per branch, and a
# footfall forecast for next month using the same actual-days-covered,
# capped-growth method as the product demand forecast.
# ---------------------------------------------------------------------------
def build_footfall(sales):
    sales = sales.copy()
    sales['Branch'] = sales['Inv.No'].apply(extract_branch)
    dt = pd.to_datetime(sales['Date'], format='mixed')
    sales['Day'] = dt.dt.date

    months = sorted(sales['Source_Month'].unique())
    latest = months[-1]
    target_month = next_month_str(latest)
    ty, tm = map(int, target_month.split('-'))
    target_days = days_in_month(ty, tm)

    daily = sales.groupby(['Day', 'Branch']).agg(Footfall=('Inv.No', 'nunique')).reset_index()
    daily = daily.sort_values('Day').reset_index(drop=True)

    monthly = sales.groupby(['Branch', 'Source_Month']).agg(Footfall=('Inv.No', 'nunique')).reset_index()
    days_covered = sales.groupby(['Branch', 'Source_Month'])['Day'].nunique().rename('Days_Covered').reset_index()
    monthly = monthly.merge(days_covered, on=['Branch', 'Source_Month'])
    monthly['Avg_Daily'] = (monthly['Footfall'] / monthly['Days_Covered']).round(2)

    rows = []
    for branch in sorted(sales['Branch'].unique()):
        sub = monthly[monthly['Branch'] == branch].set_index('Source_Month')
        series = [float(sub['Avg_Daily'].get(m, 0.0)) for m in months]
        if len(series) == 1 or series[-2] == 0:
            pred_per_day, growth = series[-1], 0.0
        else:
            growth = max(-0.3, min(0.5, (series[-1] - series[-2]) / series[-2]))
            pred_per_day = series[-1] * (1 + growth)
        pred_total = max(0, round(pred_per_day * target_days))
        rows.append((branch, series[-1] if series else 0.0, round(pred_per_day, 1), pred_total, round(growth * 100, 1)))

    forecast = pd.DataFrame(rows, columns=['Branch', 'Latest_Avg_Daily', 'Predicted_Avg_Daily',
                                            'Predicted_Total_Footfall', 'Growth_Pct'])
    return daily, monthly, forecast, target_month


# ---------------------------------------------------------------------------
# Analysis 1d: Day-by-day forecast for next month, per branch
#
# A single monthly total (build_footfall's job) doesn't say which days will
# be busy. This breaks it down by day using a day-of-week seasonality index
# learned from all 3 months of daily history (e.g. "Mondays run 15% above
# this branch's daily average"), then spreads next month's ALREADY-TRUSTED
# monthly forecast across its actual calendar days according to that
# pattern - so the daily figures always add back up to the monthly one,
# rather than being a second, disconnected guess. Same method for revenue.
#
# Deliberately business-level (per branch/day), not per-product: a
# product x day grid for 11,000+ products would be too granular to act on.
# ---------------------------------------------------------------------------
WEEKDAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


SPECIAL_DAY_LABELS = {
    'holiday': 'Holiday', 'pre_holiday': 'Day before holiday',
    'post_holiday': 'Day after holiday', 'second_saturday': '2nd Saturday',
}


def build_daywise_forecast(sales, footfall_forecast):
    sales = sales.copy()
    sales['Branch'] = sales['Inv.No'].apply(extract_branch)
    dt = pd.to_datetime(sales['Date'], format='mixed')
    sales['Day'] = dt.dt.date
    sales['Weekday'] = dt.dt.weekday

    holidays = load_holidays()
    classified = sales['Day'].apply(lambda d: classify_special_day(pd.Timestamp(d), holidays))
    sales['Day_Type'] = [c[0] for c in classified]
    sales['Day_Label'] = [c[1] for c in classified]

    months = sorted(sales['Source_Month'].unique())
    latest = months[-1]
    target_month = next_month_str(latest)
    ty, tm = map(int, target_month.split('-'))
    target_days_count = days_in_month(ty, tm)
    branches = sorted(sales['Branch'].unique())

    # dropna=False: Day_Type is None for every ordinary day - groupby() drops
    # any row whose grouping key is null by default, which would silently
    # wipe out every normal day and leave only the handful of special ones.
    daily = sales.groupby(['Day', 'Weekday', 'Branch', 'Day_Type'], dropna=False).agg(
        Footfall=('Inv.No', 'nunique'), Revenue=('Item Total', 'sum')).reset_index()

    # Weekday pattern excludes ALL special days (holidays, the day before/after
    # one, 2nd Saturdays) - none of them should skew what a "normal" Monday/
    # Tuesday/etc. looks like for that branch.
    normal_daily = daily[daily['Day_Type'].isna()]
    branch_avg = {b: {'footfall': normal_daily[normal_daily['Branch'] == b]['Footfall'].mean(),
                       'revenue': normal_daily[normal_daily['Branch'] == b]['Revenue'].mean()}
                  for b in branches}

    dow_index = {}
    for branch in branches:
        b = normal_daily[normal_daily['Branch'] == branch]
        avg_ff, avg_rev = branch_avg[branch]['footfall'], branch_avg[branch]['revenue']
        idx = {}
        for wd in range(7):
            sub = b[b['Weekday'] == wd]
            idx[wd] = {
                'footfall': float(sub['Footfall'].mean() / avg_ff) if len(sub) and avg_ff > 0 else 1.0,
                'revenue': float(sub['Revenue'].mean() / avg_rev) if len(sub) and avg_rev > 0 else 1.0,
            }
        dow_index[branch] = idx

    # For each special-day TYPE, how far its footfall/revenue actually strayed
    # (in either direction - a pre-holiday boost stays a boost) from what a
    # normal day of that same weekday would have produced. Averaged per branch,
    # falling back to the all-branch average, then to "no adjustment" if that
    # day type has never been observed yet - accuracy improves automatically as
    # more such days accumulate in the history.
    def special_ratios(day_type, metric):
        out = {}
        for _, r in daily[daily['Day_Type'] == day_type].iterrows():
            br, wd = r['Branch'], r['Weekday']
            expected = branch_avg.get(br, {}).get(metric, 0) * dow_index.get(br, {}).get(wd, {}).get(metric, 1.0)
            if expected > 0:
                out.setdefault(br, []).append(r[metric.capitalize()] / expected)
        return out

    special_factor = {}
    for day_type in SPECIAL_DAY_LABELS:
        special_factor[day_type] = {}
        for metric in ('footfall', 'revenue'):
            ratios_by_branch = special_ratios(day_type, metric)
            all_ratios = [r for rs in ratios_by_branch.values() for r in rs]
            global_avg = (sum(all_ratios) / len(all_ratios)) if all_ratios else 1.0
            for branch in branches:
                rs = ratios_by_branch.get(branch)
                special_factor[day_type].setdefault(branch, {})[metric] = (
                    sum(rs) / len(rs) if rs else global_avg)

    # Monthly revenue forecast per branch, same day-covered-adjusted trend method as footfall
    monthly_rev = sales.groupby(['Branch', 'Source_Month']).agg(Revenue=('Item Total', 'sum')).reset_index()
    days_covered = sales.groupby(['Branch', 'Source_Month'])['Day'].nunique().rename('Days_Covered').reset_index()
    monthly_rev = monthly_rev.merge(days_covered, on=['Branch', 'Source_Month'])
    monthly_rev['Avg_Daily_Rev'] = monthly_rev['Revenue'] / monthly_rev['Days_Covered']

    rev_forecast = {}
    for branch in branches:
        sub = monthly_rev[monthly_rev['Branch'] == branch].set_index('Source_Month')
        series = [float(sub['Avg_Daily_Rev'].get(m, 0.0)) for m in months]
        if len(series) == 1 or series[-2] == 0:
            pred_per_day = series[-1]
        else:
            growth = max(-0.3, min(0.5, (series[-1] - series[-2]) / series[-2]))
            pred_per_day = series[-1] * (1 + growth)
        rev_forecast[branch] = pred_per_day * target_days_count

    ff_forecast_map = footfall_forecast.set_index('Branch')['Predicted_Total_Footfall'].to_dict()
    target_dates = pd.date_range(pd.Timestamp(ty, tm, 1), periods=target_days_count, freq='D')
    target_classified = [classify_special_day(d, holidays) for d in target_dates]

    rows = []
    for branch in branches:
        w_ff, w_rev = [], []
        for d, (day_type, _) in zip(target_dates, target_classified):
            base_ff = dow_index[branch][d.weekday()]['footfall']
            base_rev = dow_index[branch][d.weekday()]['revenue']
            if day_type:
                base_ff *= special_factor[day_type][branch]['footfall']
                base_rev *= special_factor[day_type][branch]['revenue']
            w_ff.append(base_ff)
            w_rev.append(base_rev)
        sum_ff, sum_rev = sum(w_ff), sum(w_rev)
        total_ff, total_rev = ff_forecast_map.get(branch, 0), rev_forecast.get(branch, 0)
        for d, wf, wr, (day_type, label) in zip(target_dates, w_ff, w_rev, target_classified):
            rows.append({
                'Date': d.date(), 'Weekday': WEEKDAY_NAMES[d.weekday()], 'Branch': branch,
                'Holiday': label,
                'Predicted_Footfall': round(total_ff * (wf / sum_ff)) if sum_ff > 0 else 0,
                'Predicted_Revenue': round(total_rev * (wr / sum_rev), 2) if sum_rev > 0 else 0,
            })
    daywise = pd.DataFrame(rows)

    all_branches = daywise.groupby(['Date', 'Weekday', 'Holiday']).agg(
        Predicted_Footfall=('Predicted_Footfall', 'sum'), Predicted_Revenue=('Predicted_Revenue', 'sum')).reset_index()
    all_branches.insert(2, 'Branch', 'All Branches')
    daywise = pd.concat([daywise, all_branches], ignore_index=True).sort_values(['Branch', 'Date']).reset_index(drop=True)

    idx_rows = [{'Branch': br, 'Weekday': WEEKDAY_NAMES[wd],
                 'Footfall_Index': round(dow_index[br][wd]['footfall'], 2),
                 'Revenue_Index': round(dow_index[br][wd]['revenue'], 2)}
                for br in branches for wd in range(7)]
    dow_index_df = pd.DataFrame(idx_rows)

    return daywise, dow_index_df, target_month


# ---------------------------------------------------------------------------
# Analysis 1e: Employee performance - Billed By / Item Given By (sales),
# Created By (purchase). None of these columns exist in any export yet;
# this activates automatically the first time a file includes one of them.
# ---------------------------------------------------------------------------
def build_employee_performance(sales, purch, dist_lines=None):
    billed_col = find_col(sales, BILLED_BY_CANDIDATES)
    given_col = find_col(sales, GIVEN_BY_CANDIDATES)
    created_col = find_col(purch, CREATED_BY_CANDIDATES)

    result = {'billed_by': None, 'given_by': None, 'created_by': None}

    if billed_col:
        g = sales[sales[billed_col].notna()].groupby(billed_col).agg(
            Bills=('Inv.No', 'nunique'), Revenue=('Item Total', 'sum'),
            Patients=('Patient', 'nunique')).reset_index().rename(columns={billed_col: 'Employee'})
        g['Avg_Bill_Value'] = (g['Revenue'] / g['Bills']).round(2)
        result['billed_by'] = g.sort_values('Revenue', ascending=False).reset_index(drop=True)

    if given_col:
        g2 = sales[sales[given_col].notna()].groupby(given_col).agg(
            Lines=('Product', 'count'), Qty=('Qty', 'sum'), Value=('Item Total', 'sum')
        ).reset_index().rename(columns={given_col: 'Employee'})
        result['given_by'] = g2.sort_values('Value', ascending=False).reset_index(drop=True)

    if created_col:
        p = purch.copy()
        p['Factor'] = p['Factor'].replace(0, 1).fillna(1)
        # Same person gets keyed in with inconsistent casing/whitespace across
        # entries (e.g. 'NaReNdRa DeVaDiGa' / 'Narendra Devadiga' / trailing
        # spaces) - normalize before grouping so their rows don't fragment.
        p[created_col] = p[created_col].str.strip().str.title()
        g3 = p[p[created_col].notna()].groupby(created_col).agg(
            Entries=('Inv.No', 'nunique'), Lines=('Product', 'count'), Value=('Item Total', 'sum')
        ).reset_index().rename(columns={created_col: 'Employee'})
        # Cross-reference with the purchase-entry error checks already built -
        # who's actually keying in the PTR-above-MRP / MRP-mismatch entries.
        errors = p[(p['Sale Rate'] > p['MRP']) & (p['MRP'] > 0) & p[created_col].notna()]
        err_counts = errors.groupby(created_col)['Inv.No'].count().rename('PTR_Errors')
        g3 = g3.merge(err_counts, left_on='Employee', right_index=True, how='left')
        g3['PTR_Errors'] = g3['PTR_Errors'].fillna(0).astype(int)

        # Whose purchase entries carry the best/worst supplier terms - same
        # embedded-profit rule as the Distributor Scorecard (MRP-valid lines
        # only), grouped by the same normalized identity as everything else
        # above so an employee's rows stay one consistent identity across
        # every metric in this table.
        if dist_lines is not None and created_col in dist_lines.columns:
            dist_names = dist_lines[created_col].str.strip().str.title()
            pb = dist_lines.groupby(dist_names).agg(
                Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'), Embedded_Profit=('Embedded_Profit', 'sum')
            ).rename_axis('Employee')
            g3 = g3.merge(pb, left_on='Employee', right_index=True, how='left')
            g3['Max_Sell_Pretax'] = g3['Max_Sell_Pretax'].fillna(0)
            g3['Embedded_Profit'] = g3['Embedded_Profit'].fillna(0)
            g3['Margin_Pct'] = np.where(g3['Max_Sell_Pretax'] > 0, (g3['Embedded_Profit'] / g3['Max_Sell_Pretax'] * 100).round(1), 0)
            g3 = g3.drop(columns=['Max_Sell_Pretax'])

        result['created_by'] = g3.sort_values('Value', ascending=False).reset_index(drop=True)

    return result


# ---------------------------------------------------------------------------
# Analysis 1f: Customer loyalty - new / returning / dropped-off, by mobile
# number, month over month. The Mobile Number column doesn't exist in any
# export yet; this activates automatically once a file includes one.
# Patient NAME is deliberately not used as a fallback identifier - the same
# name can be different people, and it would quietly produce wrong counts
# rather than an honest "not available yet".
# ---------------------------------------------------------------------------
def build_customer_loyalty(sales):
    mobile_col = find_col(sales, MOBILE_CANDIDATES)
    if not mobile_col:
        return None, None

    s = sales[sales[mobile_col].notna()].copy()
    s[mobile_col] = s[mobile_col].astype(str).str.strip()
    months = sorted(s['Source_Month'].unique())
    monthly_customers = {m: set(s[s['Source_Month'] == m][mobile_col]) for m in months}

    rows = []
    seen_so_far = set()
    for i, m in enumerate(months):
        current = monthly_customers[m]
        new_customers = current - seen_so_far
        returning = current & seen_so_far
        prev_customers = monthly_customers[months[i - 1]] if i > 0 else set()
        dropped_off = prev_customers - current
        rows.append({
            'Month': m, 'New_Customers': len(new_customers), 'Returning_Customers': len(returning),
            'Dropped_Off_From_Prev_Month': len(dropped_off), 'Total_Active_Customers': len(current),
        })
        seen_so_far |= current

    trend = pd.DataFrame(rows)
    return trend, mobile_col


# ---------------------------------------------------------------------------
# Analysis 2: Over/under purchased (cumulative, all history)
#
# Purchase 'Qty' is counted in PACKS (Factor = units per pack, e.g. a strip of
# 10 tablets); Sales 'Qty' is counted in INDIVIDUAL UNITS. Comparing them raw
# understates purchases by ~Factor-fold. Purchase quantity must be converted
# to individual units (Qty * Factor) before it can be compared to Sales Qty,
# and the per-unit purchase rate divided by Factor for the same reason.
# ---------------------------------------------------------------------------
def build_over_under(sales, purch):
    purch = purch.copy()
    purch['Factor'] = purch['Factor'].replace(0, 1).fillna(1)
    purch['Physical_Qty'] = purch['Qty'] * purch['Factor']

    s = sales.groupby('Product').agg(Sold_Qty=('Qty', 'sum'), Sold_Value=('Item Total', 'sum')).reset_index()
    pu = purch.groupby('Product').agg(Purch_Qty=('Physical_Qty', 'sum'), Purch_Value=('Item Total', 'sum')).reset_index()
    rate = purch.sort_values('Source_Month').groupby('Product').apply(
        lambda d: (d['Item Total'].sum() / d['Physical_Qty'].sum()) if d['Physical_Qty'].sum() > 0 else np.nan,
        include_groups=False)

    m = s.merge(pu, on='Product', how='outer').fillna(0)
    m['Net_Qty'] = m['Purch_Qty'] - m['Sold_Qty']
    m['Rate'] = m['Product'].map(rate).fillna(0)
    m['Net_Value_Approx'] = (m['Net_Qty'] * m['Rate']).round(2)

    def classify(row):
        sold, p = row['Sold_Qty'], row['Purch_Qty']
        if p == 0 and sold > 0:
            return 'Sold, never purchased (old stock)'
        if sold == 0 and p > 0:
            return 'Purchased, never sold (dead stock)'
        if sold == 0 and p == 0:
            return 'No activity'
        ratio = p / sold
        if ratio >= 1.5:
            return 'Over-purchased'
        if ratio <= 0.5:
            return 'Under-purchased'
        return 'Balanced'

    m['Status'] = m.apply(classify, axis=1)
    factor_map = build_product_factor_map(sales, purch)
    for col in ('Sold_Qty', 'Purch_Qty', 'Net_Qty'):
        m[col + '_Strips'] = to_strips(m[col], m['Product'], factor_map)
    return m.sort_values('Net_Value_Approx', ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Analysis 2a: Month-over-month purchasing discipline + footfall trend
#
# Unlike build_over_under() (all history to date, cumulative), this scores
# EACH MONTH ON ITS OWN - that month's purchases vs that month's sales - so
# you can see whether buying discipline is actually improving month after
# month, not just what the all-time picture looks like.
#
# Goal directions: Over-purchased down, Under-purchased down, Balanced up,
# Dead stock down, Footfall up. Next month is projected with the same
# capped-growth trend method used for demand/footfall forecasting.
# ---------------------------------------------------------------------------
TREND_GOALS = {
    'Over_Purchased': 'down', 'Under_Purchased': 'down',
    'Balanced': 'up', 'Dead_Stock': 'down', 'Footfall': 'up',
}


def _trend_next(series):
    series = [float(x) for x in series]
    if len(series) < 2 or series[-2] == 0:
        return max(0, round(series[-1]))
    growth = max(-0.3, min(0.5, (series[-1] - series[-2]) / series[-2]))
    return max(0, round(series[-1] * (1 + growth)))


def build_monthly_trend(sales, purch, footfall_forecast):
    months = sorted(set(sales['Source_Month'].unique()) | set(purch['Source_Month'].unique()))
    rows = []
    for m in months:
        s_m = sales[sales['Source_Month'] == m]
        p_m = purch[purch['Source_Month'] == m]
        counts = {'Over_Purchased': 0, 'Under_Purchased': 0, 'Balanced': 0, 'Dead_Stock': 0}
        if not s_m.empty or not p_m.empty:
            ou = build_over_under(s_m, p_m)
            vc = ou['Status'].value_counts()
            counts = {
                'Over_Purchased': int(vc.get('Over-purchased', 0)),
                'Under_Purchased': int(vc.get('Under-purchased', 0)),
                'Balanced': int(vc.get('Balanced', 0)),
                'Dead_Stock': int(vc.get('Purchased, never sold (dead stock)', 0)),
            }
        counts['Month'] = m
        # Raw monthly footfall total is biased low for a partial (still-in-progress)
        # month - the same trap the demand/footfall forecasts already avoid. Track
        # the average-per-day rate instead, which is comparable across months
        # regardless of how many days of that month have been uploaded so far.
        counts['Footfall_Total'] = int(s_m['Inv.No'].nunique())
        days_covered = pd.to_datetime(s_m['Date'], format='mixed').dt.date.nunique() if len(s_m) else 1
        counts['Footfall_Avg_Daily'] = round(counts['Footfall_Total'] / max(1, days_covered), 1)
        rows.append(counts)

    trend = pd.DataFrame(rows)[['Month', 'Over_Purchased', 'Under_Purchased', 'Balanced', 'Dead_Stock',
                                 'Footfall_Total', 'Footfall_Avg_Daily']]

    count_metrics = ['Over_Purchased', 'Under_Purchased', 'Balanced', 'Dead_Stock']
    pred_rows = []
    for metric in count_metrics:
        series = trend[metric].tolist()
        latest, prev = series[-1], (series[-2] if len(series) >= 2 else None)
        predicted = _trend_next(series)
        goal = TREND_GOALS[metric]
        on_track = None if prev is None else (latest <= prev if goal == 'down' else latest >= prev)
        pred_rows.append({'Metric': metric, 'Goal': goal, 'Latest': latest,
                           'Previous': prev, 'Predicted_Next': predicted, 'On_Track': on_track})

    # Footfall: "on track" compares the day-normalized rate (fair across partial
    # months); the predicted total reuses build_footfall()'s own forecast, which
    # is already day-covered-adjusted and summed across branches - no need to
    # re-derive it here with a cruder method.
    avg_daily = trend['Footfall_Avg_Daily'].tolist()
    ff_prev_avg = avg_daily[-2] if len(avg_daily) >= 2 else None
    ff_on_track = None if ff_prev_avg is None else (avg_daily[-1] >= ff_prev_avg)
    pred_rows.append({'Metric': 'Footfall', 'Goal': 'up',
                       'Latest': trend['Footfall_Total'].iloc[-1],
                       'Previous': trend['Footfall_Total'].iloc[-2] if len(trend) >= 2 else None,
                       'Predicted_Next': int(footfall_forecast['Predicted_Total_Footfall'].sum()),
                       'On_Track': ff_on_track})

    prediction = pd.DataFrame(pred_rows)
    target_month = next_month_str(months[-1])
    return trend, prediction, target_month


# ---------------------------------------------------------------------------
# Analysis 2b: Gross profit & margin, product-wise
#
# Cost is computed pre-tax (GST excluded on both sides, since it's a pass-
# through - collected on sales, recoverable as input credit on purchases -
# not real income or real cost). Purchase Qty is converted from packs to
# individual units (Qty * Factor) to match Sales Qty, same fix as above.
# Products with no purchase history anywhere in the loaded data have no
# known cost and are excluded from profit totals, listed separately instead.
# ---------------------------------------------------------------------------
def build_profit_margin(sales, purch):
    purch = purch.copy()
    purch['Factor'] = purch['Factor'].replace(0, 1).fillna(1)
    purch['Physical_Qty'] = purch['Qty'] * purch['Factor']
    purch['Pretax_Value'] = purch['Qty'] * purch['Sale Rate'] - purch['Disc Amount'].fillna(0)

    cost = purch.groupby('Product').agg(
        Physical_Qty=('Physical_Qty', 'sum'), Pretax_Value=('Pretax_Value', 'sum')).reset_index()
    cost = cost[cost['Physical_Qty'] > 0]
    cost['Cost_Per_Unit'] = (cost['Pretax_Value'] / cost['Physical_Qty']).round(4)
    cost_map = cost.set_index('Product')['Cost_Per_Unit']

    sales = sales.copy()
    sales['Pretax_Value'] = sales['Item Total'] / (1 + sales['Tax Rate'].fillna(0) / 100)

    known = sales[sales['Product'].isin(cost_map.index)].copy()
    unknown = sales[~sales['Product'].isin(cost_map.index)].copy()

    g = known.groupby('Product').agg(
        Qty_Sold=('Qty', 'sum'), Revenue=('Item Total', 'sum'), Pretax_Revenue=('Pretax_Value', 'sum')).reset_index()
    g['Cost_Per_Unit'] = g['Product'].map(cost_map)
    g['COGS'] = (g['Qty_Sold'] * g['Cost_Per_Unit']).round(2)
    g['Gross_Profit'] = (g['Pretax_Revenue'] - g['COGS']).round(2)
    g['Margin_Pct'] = np.where(g['Pretax_Revenue'] > 0, (g['Gross_Profit'] / g['Pretax_Revenue'] * 100).round(1), 0)
    factor_map = build_product_factor_map(sales, purch)
    g['Qty_Sold_Strips'] = to_strips(g['Qty_Sold'], g['Product'], factor_map)
    g = g.sort_values('Gross_Profit', ascending=False).reset_index(drop=True)

    unk = unknown.groupby('Product').agg(Qty_Sold=('Qty', 'sum'), Revenue=('Item Total', 'sum')).reset_index()
    unk['Qty_Sold_Strips'] = to_strips(unk['Qty_Sold'], unk['Product'], factor_map)
    unk = unk.sort_values('Revenue', ascending=False).reset_index(drop=True)

    return g, unk


# ---------------------------------------------------------------------------
# Analysis 3: Purchase entry errors (PTR>MRP, MRP missing, MRP variance) on latest month
# ---------------------------------------------------------------------------
def build_purchase_errors(purch, latest_month):
    latest = purch[purch['Source_Month'] == latest_month].copy()

    ptr_high = latest[(latest['Sale Rate'] > latest['MRP']) & (latest['MRP'] > 0)].copy()
    ptr_high['Excess'] = (ptr_high['Sale Rate'] - ptr_high['MRP']).round(2)

    gifts = latest[(latest['MRP'] <= 0) & (latest['Sale Rate'] <= 1)].copy()
    mrp_missing = latest[(latest['MRP'] <= 0) & (latest['Sale Rate'] > 1)].copy()

    g = purch.groupby('Product')['MRP'].agg(['min', 'max', 'count']).reset_index()
    g = g[g['count'] >= 2]
    g['ratio'] = g['max'] / g['min'].replace(0, np.nan)
    variance_products = g[(g['ratio'] >= 3) & (g['min'] > 0)].sort_values('ratio', ascending=False)

    return ptr_high, mrp_missing, gifts, variance_products


# ---------------------------------------------------------------------------
# Analysis 4: Scheme consistency (10+1, 10+2 style free-qty offers)
# ---------------------------------------------------------------------------
def build_scheme_consistency(purch, latest_month):
    hist = purch[purch['Source_Month'] != latest_month]
    latest = purch[purch['Source_Month'] == latest_month].copy()
    latest['FreeRatio'] = np.where(latest['Qty'] > 0, latest['Free Qty'].fillna(0) / latest['Qty'], np.nan)

    scheme_hist = hist[hist['Free Qty'].fillna(0) > 0].copy()
    scheme_hist['FreeRatio'] = np.where(scheme_hist['Qty'] > 0,
                                         scheme_hist['Free Qty'] / scheme_hist['Qty'], np.nan)
    baseline = scheme_hist.groupby(['Product', 'Supplier']).agg(
        Typical_Ratio=('FreeRatio', 'median'),
        Best_Ratio=('FreeRatio', 'max'),
        Times_Seen=('FreeRatio', 'count')).reset_index()
    baseline = baseline[baseline['Times_Seen'] >= 2]

    check = latest.merge(baseline, on=['Product', 'Supplier'], how='inner')
    missed = check[(check['FreeRatio'].fillna(0)) < (check['Typical_Ratio'] * 0.8)].copy()
    missed['Shortfall_Qty'] = ((missed['Typical_Ratio'] * missed['Qty']) - missed['Free Qty'].fillna(0)).round(1)
    missed = missed.sort_values('Shortfall_Qty', ascending=False)
    return missed, baseline


# ---------------------------------------------------------------------------
# Analysis 5: Discount consistency
# ---------------------------------------------------------------------------
def build_discount_consistency(purch, latest_month):
    hist = purch[purch['Source_Month'] != latest_month]
    latest = purch[purch['Source_Month'] == latest_month].copy()

    baseline = hist[hist['Disc Percentage'] > 0].groupby(['Product', 'Supplier']).agg(
        Typical_Disc=('Disc Percentage', 'median'),
        Best_Disc=('Disc Percentage', 'max'),
        Times_Seen=('Disc Percentage', 'count')).reset_index()
    baseline = baseline[baseline['Times_Seen'] >= 2]

    check = latest.merge(baseline, on=['Product', 'Supplier'], how='inner')
    missed = check[check['Disc Percentage'] < (check['Typical_Disc'] * 0.8)].copy()
    missed['Disc_Gap_pct_pts'] = (missed['Typical_Disc'] - missed['Disc Percentage']).round(2)
    missed['Value_Lost_Approx'] = (missed['Item Total'] / (1 - missed['Disc Percentage'] / 100 + 1e-9)
                                    * (missed['Typical_Disc'] - missed['Disc Percentage']) / 100).round(2)
    missed = missed.sort_values('Value_Lost_Approx', ascending=False)
    return missed, baseline


# ---------------------------------------------------------------------------
# Analysis 6: Distributor (Supplier) Profit Scorecard
#
# For each distributor, the POTENTIAL/embedded margin locked into what we
# bought from them: if every unit received (including free-scheme goods)
# were sold at MRP, what would that be worth versus what we actually paid.
# A buying/negotiation KPI ("who gives the best terms") - NOT realized P&L,
# since most of this stock is still unsold.
#
# Pre-tax on both sides (MRP is tax-inclusive, so GST is stripped out before
# comparing). Bill Disc (invoice-level) is deliberately ignored - only
# line-level PTR (Sale Rate) and Disc Amount count as cost. Lines with
# MRP<=0 can't be valued and are excluded from every rollup below.
# ---------------------------------------------------------------------------
TOP_N_PRODUCTS_PER_DISTRIBUTOR = 15


def compute_distributor_lines(purch):
    p = purch.copy()
    # Inv.No mixes plain ints, leading-zero codes, and slash-suffixed formats
    # ('1274/26-27') across suppliers - force to string so grouping/counting
    # by invoice can't silently split or miss rows over a dtype mismatch.
    p['Inv.No'] = p['Inv.No'].astype(str)
    p['Free Qty'] = p['Free Qty'].fillna(0)
    p['Disc Amount'] = p['Disc Amount'].fillna(0)
    p['Tax Rate'] = p['Tax Rate'].fillna(0)

    # "MRP <= 0" and "MRP > 0" look complementary but aren't - a blank (NaN)
    # MRP cell satisfies neither, so it would silently vanish from both
    # excluded and lines. Splitting on valid_mrp/~valid_mrp instead guarantees
    # every row lands in exactly one bucket.
    valid_mrp = p['MRP'] > 0
    excluded = p[~valid_mrp]
    lines = p[valid_mrp].copy()
    lines['Units_Received'] = lines['Qty'] + lines['Free Qty']
    lines['Max_Sell_Pretax'] = (lines['MRP'] / (1 + lines['Tax Rate'] / 100)) * lines['Units_Received']
    lines['Net_Cost_Pretax'] = lines['Qty'] * lines['Sale Rate'] - lines['Disc Amount']
    lines['Embedded_Profit'] = lines['Max_Sell_Pretax'] - lines['Net_Cost_Pretax']
    return lines, excluded


def build_distributor_summary(lines):
    g = lines.groupby('Supplier').agg(
        Invoices=('Inv.No', 'nunique'), Lines=('Product', 'count'),
        Total_Invoice_Value=('Item Total', 'sum'),
        Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'), Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'),
        Embedded_Profit=('Embedded_Profit', 'sum'), Months_Active=('Source_Month', 'nunique')).reset_index()
    g['Margin_Pct'] = np.where(g['Max_Sell_Pretax'] > 0, (g['Embedded_Profit'] / g['Max_Sell_Pretax'] * 100).round(1), 0)
    return g.sort_values('Embedded_Profit', ascending=False).reset_index(drop=True)


def build_distributor_month(lines):
    g = lines.groupby(['Supplier', 'Source_Month']).agg(
        Invoice_Value=('Item Total', 'sum'), Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'), Embedded_Profit=('Embedded_Profit', 'sum')).reset_index()
    g['Margin_Pct'] = np.where(g['Max_Sell_Pretax'] > 0, (g['Embedded_Profit'] / g['Max_Sell_Pretax'] * 100).round(1), 0)
    supplier_order = lines.groupby('Supplier')['Embedded_Profit'].sum().sort_values(ascending=False).index.tolist()
    g['Supplier'] = pd.Categorical(g['Supplier'], categories=supplier_order, ordered=True)
    return g.sort_values(['Supplier', 'Source_Month']).reset_index(drop=True), supplier_order


def build_distributor_per_invoice(lines):
    g = lines.groupby(['Inv.No', 'Supplier']).agg(
        Date=('Date', 'first'), Source_Month=('Source_Month', 'first'), Lines=('Product', 'count'),
        Invoice_Value=('Item Total', 'sum'), Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'), Embedded_Profit=('Embedded_Profit', 'sum')).reset_index()
    g['Margin_Pct'] = np.where(g['Max_Sell_Pretax'] > 0, (g['Embedded_Profit'] / g['Max_Sell_Pretax'] * 100).round(1), 0)
    return g.sort_values('Embedded_Profit', ascending=False).reset_index(drop=True)


def build_distributor_top_products(lines, supplier_order, top_n=TOP_N_PRODUCTS_PER_DISTRIBUTOR):
    g = lines.groupby(['Supplier', 'Product']).agg(
        Units_Received=('Units_Received', 'sum'), Max_Sell_Pretax=('Max_Sell_Pretax', 'sum'),
        Net_Cost_Pretax=('Net_Cost_Pretax', 'sum'), Embedded_Profit=('Embedded_Profit', 'sum')).reset_index()
    g['Margin_Pct'] = np.where(g['Max_Sell_Pretax'] > 0, (g['Embedded_Profit'] / g['Max_Sell_Pretax'] * 100).round(1), 0)
    parts = [g[g['Supplier'] == s].sort_values('Embedded_Profit', ascending=False).head(top_n) for s in supplier_order]
    return pd.concat(parts, ignore_index=True) if parts else g.iloc[0:0]


# ---------------------------------------------------------------------------
# Excel report builder
# ---------------------------------------------------------------------------
FONT = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=11)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color='555555')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


_table_counter = [0]


def write_df(ws, df, start_row=1, money_cols=None, qty_cols=None, highlight_col=None, highlight_map=None):
    money_cols = money_cols or []
    qty_cols = qty_cols or []
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            colname = df.columns[j - 1]
            if colname in money_cols:
                c.number_format = '#,##0.00'
            if colname in qty_cols:
                c.number_format = '#,##0'
            if highlight_col and colname == highlight_col and highlight_map:
                fill = highlight_map.get(val)
                if fill:
                    c.fill = fill
    last_row = start_row + len(df)
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        _table_counter[0] += 1
        tab = Table(displayName=f'T{_table_counter[0]}', ref=f'A{start_row}:{last_col}{last_row}')
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
        ws.add_table(tab)
    return last_row


def autosize(ws, widths):
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build_report(sales, purch, out_path):
    forecast, target_month, all_months = build_demand_forecast(sales)
    branch_summary, branch_forecast = build_branch_report(sales)
    footfall_daily, footfall_monthly, footfall_forecast, footfall_target_month = build_footfall(sales)
    dist_lines, dist_excluded = compute_distributor_lines(purch)
    employee_perf = build_employee_performance(sales, purch, dist_lines)
    customer_loyalty, mobile_col = build_customer_loyalty(sales)
    monthly_trend, trend_prediction, trend_target_month = build_monthly_trend(sales, purch, footfall_forecast)
    daywise_forecast, dow_index, daywise_target_month = build_daywise_forecast(sales, footfall_forecast)
    over_under = build_over_under(sales, purch)
    profit, profit_unknown = build_profit_margin(sales, purch)
    latest_month = all_months[-1]
    ptr_high, mrp_missing, gifts, variance = build_purchase_errors(purch, latest_month)
    scheme_missed, scheme_baseline = build_scheme_consistency(purch, latest_month)
    disc_missed, disc_baseline = build_discount_consistency(purch, latest_month)
    dist_summary = build_distributor_summary(dist_lines)
    dist_month, dist_supplier_order = build_distributor_month(dist_lines)
    dist_per_invoice = build_distributor_per_invoice(dist_lines)
    dist_top_products = build_distributor_top_products(dist_lines, dist_supplier_order)

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    ws['B2'] = 'Pharmacy Ready Reckoner'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = f'History loaded: {", ".join(all_months)}  |  Forecasting: {target_month}  |  Latest month checked for issues: {latest_month}'
    ws['B3'].font = SUBTITLE_FONT

    r = 5
    stats = [
        ('Months of history', len(all_months)),
        (f'Demand forecast built for {target_month}', f'{len(forecast)} products'),
        (f'Predicted sales value, {target_month}', round(forecast["Predicted_Value"].sum(), 2)),
        ('Branches detected', ', '.join(branch_summary['Branch'].tolist())),
        (f'Predicted total footfall, {footfall_target_month} (all branches)',
         int(footfall_forecast['Predicted_Total_Footfall'].sum())),
        ('Over-purchased products (all-time)', int((over_under['Status'] == 'Over-purchased').sum())),
        ('Dead stock products (purchased, never sold)', int((over_under['Status'] == 'Purchased, never sold (dead stock)').sum())),
        ('Gross profit, all history (pre-tax, known-cost products)', round(profit['Gross_Profit'].sum(), 2)),
        ('Overall gross margin %, known-cost products', round(profit['Gross_Profit'].sum() / profit['Pretax_Revenue'].sum() * 100, 1) if profit['Pretax_Revenue'].sum() > 0 else 0),
        ('Revenue with cost unknown (excluded from profit above)', round(profit_unknown['Revenue'].sum(), 2)),
        (f'PTR-higher-than-MRP entries in {latest_month}', len(ptr_high)),
        (f'MRP missing entries in {latest_month}', len(mrp_missing)),
        ('Products with MRP varying 3x+ (all-time)', len(variance)),
        (f'Scheme (10+1 style) shortfalls in {latest_month}', len(scheme_missed)),
        (f'Discount shortfalls in {latest_month}', len(disc_missed)),
        ('Distributors analyzed for embedded profit (purchase-side)', len(dist_summary)),
        ('Top distributor by embedded profit (all-history)',
         f"{dist_summary.iloc[0]['Supplier']} ({dist_summary.iloc[0]['Margin_Pct']}% margin)" if len(dist_summary) else 'n/a'),
        ('Purchase lines excluded from Distributor Scorecard (MRP missing)', len(dist_excluded)),
    ]
    for label, val in stats:
        ws.cell(row=r, column=2, value=label).font = BODY_FONT
        c = ws.cell(row=r, column=5, value=val)
        c.font = Font(name=FONT, bold=True, size=10)
        r += 1

    r += 2
    ws.cell(row=r, column=2, value='Tabs in this report').font = Font(name=FONT, bold=True, size=11)
    r += 1
    tab_notes = [
        'Demand Forecast - every product, predicted quantity & value for next month, trend flag.',
        'Branch-wise - revenue and growth per branch (from the invoice number prefix), plus a per-branch product forecast. Sales only - purchase invoices carry no branch code.',
        'Footfall - daily unique-bill count per branch (footfall), charted, with a next-month footfall forecast per branch.',
        'Monthly Trends - purchasing discipline (over/under-purchased, balanced, dead stock) and footfall, month by month, against goals - with next-month predictions.',
        'Day-wise Forecast - next month broken down by calendar day per branch (footfall & revenue), using each branch\'s day-of-week pattern from all 3 months of history.',
        'Employee Performance - Billed By / Item Given By / Created By breakdown. Activates once your export includes one of these columns.',
        'Customer Loyalty - new/returning/dropped-off customers by mobile number, month by month. Activates once your export includes a mobile number column.',
        'Over-Purchased - bought well more than sold (but still sold some), biggest excess value first.',
        'Dead Stock - bought but never sold at all, separate from Over-Purchased, highest value tied up first.',
        'Under-Purchased - sold well more than bought, biggest shortfall value first.',
        'Other Purchase Status - balanced, old-stock, and no-activity products, for reference.',
        'Profit & Margin - gross profit and margin % per product, pre-tax, all history. Products with no purchase record are listed separately (cost unknown).',
        'PTR Higher Than MRP - purchase rate above MRP this month - fix these entries.',
        'MRP Issues (missing/variance) - MRP=0 or same product priced very differently across bills.',
        'Scheme Shortfall - product+supplier pairs that normally get a free-qty scheme (10+1 etc.) but got less/none this month.',
        'Discount Shortfall - product+supplier pairs getting a noticeably lower discount % than their usual history.',
        'Distributor Summary - potential profit margin embedded in what each distributor supplied, all history (max sellable at MRP vs what we paid, free-scheme goods included). A buying/negotiation KPI, NOT realized profit - most stock is still unsold.',
        'Distributor x Month - the same, broken down by month per distributor, plus a supplier x month pivot.',
        'Distributor Per Invoice - potential margin for each individual purchase invoice.',
        f'Top Products by Distributor - each distributor\'s best-margin products, capped at top {TOP_N_PRODUCTS_PER_DISTRIBUTOR} per distributor.',
    ]
    for n in tab_notes:
        ws.cell(row=r, column=2, value='- ' + n).font = Font(name=FONT, size=9, italic=True, color='555555')
        r += 1

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 2
    ws.column_dimensions['D'].width = 2
    ws.column_dimensions['E'].width = 16

    # ---- Demand Forecast ----
    ws = wb.create_sheet('Demand Forecast')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Qty shown in strips (Qty / Factor) - the pack size, not individual tablets/ml'
    ws['A1'].font = Font(name=FONT, italic=True, size=9, color='555555')
    df1 = forecast.rename(columns={'Predicted_Qty_Strips': f'Predicted Qty ({target_month}, strips)',
                                    'Avg_Price': 'Avg Selling Price',
                                    'Predicted_Value': f'Predicted Value ({target_month})'})
    df1 = df1[['Product', 'Trend', f'Predicted Qty ({target_month}, strips)', 'Avg Selling Price', f'Predicted Value ({target_month})']]
    write_df(ws, df1, start_row=2, money_cols=['Avg Selling Price', f'Predicted Value ({target_month})'])
    autosize(ws, [38, 30, 20, 16, 18])
    ws.freeze_panes = 'A3'

    # ---- Branch-wise ----
    ws = wb.create_sheet('Branch-wise')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Branch summary (all history) - identified from the Inv.No prefix; sales only, purchases carry no branch code'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    month_rev_cols = [f'Revenue_{m}' for m in all_months]
    bs = branch_summary.rename(columns={'Total_Revenue': 'Total Revenue', 'Growth_Pct': f'Growth % ({all_months[0]} to {all_months[-1]})',
                                         **{c: c.replace('Revenue_', 'Revenue ') for c in month_rev_cols}})
    bs_cols = ['Branch', 'Total Revenue'] + [c.replace('Revenue_', 'Revenue ') for c in month_rev_cols] + \
              ['Invoices', 'Patients', f'Growth % ({all_months[0]} to {all_months[-1]})']
    bs = bs[bs_cols]
    last_b = write_df(ws, bs, start_row=2, money_cols=['Total Revenue'] + [c.replace('Revenue_', 'Revenue ') for c in month_rev_cols],
                       qty_cols=['Invoices', 'Patients'])

    r4 = last_b + 3
    ws.cell(row=r4, column=1, value=f'Per-branch product forecast for {target_month} (qty in strips)').font = Font(name=FONT, bold=True, size=11)
    bf = branch_forecast.rename(columns={'Predicted_Qty_Strips': 'Predicted Qty (strips)', 'Avg_Price': 'Avg Selling Price',
                                          'Predicted_Value': 'Predicted Value'})
    bf = bf[['Branch', 'Product', 'Trend', 'Predicted Qty (strips)', 'Avg Selling Price', 'Predicted Value']]
    write_df(ws, bf, start_row=r4 + 1, money_cols=['Avg Selling Price', 'Predicted Value'])
    autosize(ws, [16, 38, 16, 16, 16, 16])
    ws.freeze_panes = 'A3'

    # ---- Footfall ----
    ws = wb.create_sheet('Footfall')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Daily footfall (unique bills per day) by branch - "bills on a date" = that day\'s footfall'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)

    branch_order = branch_summary['Branch'].tolist()
    pivot = footfall_daily.pivot(index='Day', columns='Branch', values='Footfall').fillna(0)
    pivot = pivot.reindex(columns=branch_order).reset_index()

    start_row = 2
    write_df(ws, pivot, start_row=start_row, qty_cols=branch_order)
    data_last_row = start_row + len(pivot)

    chart = LineChart()
    chart.title = 'Daily footfall by branch'
    chart.style = 2
    chart.y_axis.title = 'Footfall (unique bills)'
    chart.x_axis.title = 'Date'
    chart.height = 9
    chart.width = 22
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=data_last_row)
    for i in range(2, 2 + len(branch_order)):
        data_ref = Reference(ws, min_col=i, min_row=start_row, max_row=data_last_row)
        chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f'{get_column_letter(len(branch_order) + 3)}{start_row}')

    r5 = data_last_row + 3
    ws.cell(row=r5, column=1, value=f'Footfall forecast for {footfall_target_month}').font = Font(name=FONT, bold=True, size=11)
    ff = footfall_forecast.rename(columns={'Latest_Avg_Daily': f'Avg Daily ({all_months[-1]})',
                                            'Predicted_Avg_Daily': 'Predicted Avg Daily',
                                            'Predicted_Total_Footfall': 'Predicted Total Footfall',
                                            'Growth_Pct': 'Growth %'})
    write_df(ws, ff, start_row=r5 + 1,
             money_cols=[f'Avg Daily ({all_months[-1]})', 'Predicted Avg Daily', 'Growth %'],
             qty_cols=['Predicted Total Footfall'])
    autosize(ws, [16] + [14] * len(branch_order))
    ws.freeze_panes = 'A3'

    # ---- Monthly Trends ----
    ws = wb.create_sheet('Monthly Trends')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Purchasing discipline and footfall, month by month - each month scored on its OWN purchases vs sales (not cumulative)'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    ws['A2'] = 'Goal: Over-Purchased down, Under-Purchased down, Balanced up, Dead Stock down, Footfall up'
    ws['A2'].font = Font(name=FONT, italic=True, size=9, color='555555')

    mt = monthly_trend.rename(columns={'Over_Purchased': 'Over-Purchased', 'Under_Purchased': 'Under-Purchased',
                                        'Dead_Stock': 'Dead Stock', 'Footfall_Total': 'Footfall (total bills)',
                                        'Footfall_Avg_Daily': 'Footfall (avg/day)'})
    t_start = 4
    write_df(ws, mt, start_row=t_start,
             qty_cols=['Over-Purchased', 'Under-Purchased', 'Balanced', 'Dead Stock', 'Footfall (total bills)'],
             money_cols=['Footfall (avg/day)'])
    t_last = t_start + len(mt)

    chart1 = LineChart()
    chart1.title = 'Purchasing discipline, month by month (product count)'
    chart1.style = 2
    chart1.y_axis.title = 'Products'
    chart1.x_axis.title = 'Month'
    chart1.height = 9
    chart1.width = 16
    cats1 = Reference(ws, min_col=1, min_row=t_start + 1, max_row=t_last)
    for col in range(2, 6):  # Over-Purchased, Under-Purchased, Balanced, Dead Stock
        chart1.add_data(Reference(ws, min_col=col, min_row=t_start, max_row=t_last), titles_from_data=True)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, f'H{t_start}')

    chart2 = LineChart()
    chart2.title = 'Footfall, avg bills/day (day-covered adjusted)'
    chart2.style = 10
    chart2.y_axis.title = 'Avg bills/day'
    chart2.x_axis.title = 'Month'
    chart2.height = 9
    chart2.width = 16
    cats2 = Reference(ws, min_col=1, min_row=t_start + 1, max_row=t_last)
    chart2.add_data(Reference(ws, min_col=7, min_row=t_start, max_row=t_last), titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, f'H{t_start + 20}')

    r6 = t_last + 3
    ws.cell(row=r6, column=1, value=f'Prediction for {trend_target_month}').font = Font(name=FONT, bold=True, size=11)
    tp = trend_prediction.copy()
    tp['Metric'] = tp['Metric'].str.replace('_', ' ')
    tp['On_Track'] = tp['On_Track'].map({True: 'Yes - improving', False: 'No - check this', None: 'n/a (first month)'})
    tp = tp.rename(columns={'Goal': 'Goal Direction', 'Predicted_Next': 'Predicted Next Month', 'On_Track': 'On Track?'})
    ON_TRACK_COLORS = {'Yes - improving': PatternFill('solid', fgColor='C6EFCE'),
                        'No - check this': PatternFill('solid', fgColor='FFC7CE'),
                        'n/a (first month)': PatternFill('solid', fgColor='F2F2F2')}
    write_df(ws, tp, start_row=r6 + 1, qty_cols=['Latest', 'Previous', 'Predicted Next Month'],
             highlight_col='On Track?', highlight_map=ON_TRACK_COLORS)
    autosize(ws, [18, 14, 20, 24, 24, 12, 12, 12])
    ws.freeze_panes = 'A5'

    # ---- Day-wise Forecast ----
    ws = wb.create_sheet('Day-wise Forecast')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f'{daywise_target_month} forecast by calendar day per branch - each branch\'s next-month total spread across its days using its own day-of-week pattern from all 3 months of history'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)

    ws['A2'] = 'Day-of-week pattern (index = that weekday\'s average vs. the branch\'s overall daily average - 1.20 means 20% busier than average)'
    ws['A2'].font = Font(name=FONT, italic=True, size=9, color='555555')
    dow_out = dow_index.rename(columns={'Footfall_Index': 'Footfall Index', 'Revenue_Index': 'Revenue Index'})
    idx_last = write_df(ws, dow_out, start_row=4, money_cols=['Footfall Index', 'Revenue Index'])

    r7 = idx_last + 3
    ws.cell(row=r7, column=1, value=f'Day-by-day forecast for {daywise_target_month} (holiday-adjusted where a declared holiday falls, see holidays.json)').font = Font(name=FONT, bold=True, size=11)
    dw_out = daywise_forecast.rename(columns={'Predicted_Footfall': 'Predicted Footfall', 'Predicted_Revenue': 'Predicted Revenue'})
    dw_out = dw_out[['Date', 'Weekday', 'Branch', 'Holiday', 'Predicted Footfall', 'Predicted Revenue']]
    dw_start = r7 + 1
    write_df(ws, dw_out, start_row=dw_start, qty_cols=['Predicted Footfall'], money_cols=['Predicted Revenue'])
    dw_last = dw_start + len(dw_out)

    holiday_fill = PatternFill('solid', fgColor='F3E6CB')
    for i, is_holiday in enumerate(dw_out['Holiday'] != ''):
        if is_holiday:
            for col in range(1, len(dw_out.columns) + 1):
                ws.cell(row=dw_start + 1 + i, column=col).fill = holiday_fill

    # 'All Branches' sorts first alphabetically (A < H < S), so it occupies the
    # first `all_count` data rows of the table just written.
    all_count = int((daywise_forecast['Branch'] == 'All Branches').sum())
    all_first_row, all_last_row = dw_start + 1, dw_start + all_count

    chart = LineChart()
    chart.title = f'{daywise_target_month} predicted footfall by day, all branches'
    chart.style = 12
    chart.y_axis.title = 'Predicted footfall'
    chart.x_axis.title = 'Date'
    chart.height = 9
    chart.width = 22
    date_col_idx = dw_out.columns.get_loc('Date') + 1
    ff_col_idx = dw_out.columns.get_loc('Predicted Footfall') + 1
    cats = Reference(ws, min_col=date_col_idx, min_row=all_first_row, max_row=all_last_row)
    chart.add_data(Reference(ws, min_col=ff_col_idx, min_row=dw_start, max_row=all_last_row), titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f'H{dw_start}')

    autosize(ws, [14, 12, 16, 18, 18, 18])
    ws.freeze_panes = 'A5'

    # ---- Employee Performance ----
    ws = wb.create_sheet('Employee Performance')
    ws.sheet_view.showGridLines = False
    r_emp = 2
    any_emp_data = any(v is not None for v in employee_perf.values())
    if not any_emp_data:
        ws['A1'] = 'Not available yet'
        ws['A1'].font = Font(name=FONT, bold=True, size=13)
        ws['A2'] = 'None of your exports include a "Billed By", "Item Given By", or "Created By" column yet.'
        ws['A2'].font = Font(name=FONT, size=10, color='555555')
        ws['A3'] = 'This activates automatically the first time an uploaded file includes one of these.'
        ws['A3'].font = Font(name=FONT, size=10, color='555555')
    else:
        if employee_perf['billed_by'] is not None:
            ws.cell(row=r_emp, column=1, value='Billed By (sales) - revenue, bills, and average bill value per employee').font = Font(name=FONT, bold=True, size=11)
            df_b = employee_perf['billed_by'].rename(columns={'Bills': 'Bills', 'Revenue': 'Revenue',
                                                                'Patients': 'Unique Patients', 'Avg_Bill_Value': 'Avg Bill Value'})
            r_emp = write_df(ws, df_b, start_row=r_emp + 1, qty_cols=['Bills', 'Unique Patients'],
                              money_cols=['Revenue', 'Avg Bill Value']) + 3
        if employee_perf['given_by'] is not None:
            ws.cell(row=r_emp, column=1, value='Item Given By (sales) - lines dispensed, quantity, and value per employee').font = Font(name=FONT, bold=True, size=11)
            df_g = employee_perf['given_by'].rename(columns={'Lines': 'Line Items', 'Qty': 'Total Qty', 'Value': 'Total Value'})
            r_emp = write_df(ws, df_g, start_row=r_emp + 1, qty_cols=['Line Items', 'Total Qty'], money_cols=['Total Value']) + 3
        if employee_perf['created_by'] is not None:
            ws.cell(row=r_emp, column=1,
                    value='Created By (purchase entries) - cross-checked against PTR-above-MRP errors, and against '
                          'the embedded profit/margin those entries carry (same rule as the Distributor Scorecard, '
                          'MRP-valid lines only) - whose entries land the best/worst supplier terms').font = Font(name=FONT, bold=True, size=11)
            df_c = employee_perf['created_by'].rename(columns={'Entries': 'Invoices Entered', 'Lines': 'Line Items',
                                                                 'Value': 'Total Purchase Value', 'PTR_Errors': 'PTR > MRP Errors',
                                                                 'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
            has_profit_cols = 'Embedded Profit' in df_c.columns
            money_cols = ['Total Purchase Value'] + (['Embedded Profit', 'Margin %'] if has_profit_cols else [])
            write_df(ws, df_c, start_row=r_emp + 1, qty_cols=['Invoices Entered', 'Line Items', 'PTR > MRP Errors'],
                     money_cols=money_cols)
    autosize(ws, [26, 14, 16, 16, 16, 16, 10])
    ws.freeze_panes = 'A2'

    # ---- Customer Loyalty ----
    ws = wb.create_sheet('Customer Loyalty')
    ws.sheet_view.showGridLines = False
    if customer_loyalty is None:
        ws['A1'] = 'Not available yet'
        ws['A1'].font = Font(name=FONT, bold=True, size=13)
        ws['A2'] = 'None of your exports include a mobile number column yet.'
        ws['A2'].font = Font(name=FONT, size=10, color='555555')
        ws['A3'] = 'This activates automatically the first time an uploaded file includes one - identifying customers by name alone isn\'t reliable, since names repeat across different people.'
        ws['A3'].font = Font(name=FONT, size=10, color='555555')
    else:
        ws['A1'] = f'New / returning / dropped-off customers by month, identified by "{mobile_col}". Goal: New and Returning up, Dropped Off down.'
        ws['A1'].font = Font(name=FONT, bold=True, size=11)
        cl = customer_loyalty.rename(columns={'New_Customers': 'New Customers', 'Returning_Customers': 'Returning Customers',
                                               'Dropped_Off_From_Prev_Month': 'Dropped Off (vs prev month)',
                                               'Total_Active_Customers': 'Total Active This Month'})
        write_df(ws, cl, start_row=3, qty_cols=['New Customers', 'Returning Customers',
                                                 'Dropped Off (vs prev month)', 'Total Active This Month'])
    autosize(ws, [14, 16, 18, 22, 20])
    ws.freeze_panes = 'A4'

    # ---- Over-Purchased / Under-Purchased (separated, each sorted by value) ----
    # Qty columns are in strips (Qty / Factor), not individual tablets/ml.
    ou = over_under.rename(columns={'Sold_Qty_Strips': 'Total Sold Qty (strips)', 'Purch_Qty_Strips': 'Total Purchased Qty (strips)',
                                     'Net_Qty_Strips': 'Net Qty (strips)', 'Sold_Value': 'Total Sale Value',
                                     'Purch_Value': 'Total Purchase Value'})
    base_cols = ['Product', 'Total Purchased Qty (strips)', 'Total Sold Qty (strips)', 'Net Qty (strips)', 'Rate',
                 'Total Purchase Value', 'Total Sale Value']

    over_df = ou[ou['Status'] == 'Over-purchased'].copy()
    over_df = over_df.rename(columns={'Net_Value_Approx': 'Excess Value (at cost)'})
    over_df = over_df.sort_values('Excess Value (at cost)', ascending=False)
    over_df = over_df[base_cols + ['Excess Value (at cost)']]

    dead_df = ou[ou['Status'] == 'Purchased, never sold (dead stock)'].copy()
    dead_df = dead_df.rename(columns={'Net_Value_Approx': 'Value Tied Up (at cost)'})
    dead_df = dead_df.sort_values('Value Tied Up (at cost)', ascending=False)
    dead_df = dead_df[base_cols + ['Value Tied Up (at cost)']]

    under_df = ou[ou['Status'] == 'Under-purchased'].copy()
    under_df['Shortfall Value (at cost)'] = -under_df['Net_Value_Approx']
    under_df = under_df.sort_values('Shortfall Value (at cost)', ascending=False)
    under_df = under_df[base_cols + ['Shortfall Value (at cost)']]

    other_df = ou[ou['Status'].isin(['Balanced', 'Sold, never purchased (old stock)', 'No activity'])].copy()
    other_df = other_df.rename(columns={'Net_Value_Approx': 'Net Value (at cost)'})
    other_df = other_df.sort_values('Total Sale Value', ascending=False)
    other_df = other_df[base_cols[:4] + ['Status'] + base_cols[4:] + ['Net Value (at cost)']]

    ws = wb.create_sheet('Over-Purchased')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Bought at least 1.5x what was sold (and it did sell at least some) - biggest excess first. Qty in strips.'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    write_df(ws, over_df, start_row=2, money_cols=['Rate', 'Total Purchase Value', 'Total Sale Value', 'Excess Value (at cost)'])
    autosize(ws, [38, 18, 16, 14, 12, 16, 16, 18])
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Dead Stock')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Purchased in this history but never sold at all - zero movement, highest value tied up first. Qty in strips.'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    write_df(ws, dead_df, start_row=2, money_cols=['Rate', 'Total Purchase Value', 'Total Sale Value', 'Value Tied Up (at cost)'])
    autosize(ws, [38, 18, 16, 14, 12, 16, 16, 18])
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Under-Purchased')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Sold at least 2x what was purchased - biggest shortfall first (likely running down stock bought before this history). Qty in strips.'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    write_df(ws, under_df, start_row=2, money_cols=['Rate', 'Total Purchase Value', 'Total Sale Value', 'Shortfall Value (at cost)'])
    autosize(ws, [38, 18, 16, 14, 12, 16, 16, 18])
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Other Purchase Status')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Balanced, sold-with-no-purchase-record (old stock), and no-activity products - sorted by sale value. Qty in strips.'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    write_df(ws, other_df, start_row=2, money_cols=['Rate', 'Total Purchase Value', 'Total Sale Value', 'Net Value (at cost)'])
    autosize(ws, [38, 18, 16, 14, 26, 12, 16, 16, 18])
    ws.freeze_panes = 'A3'

    # ---- Profit & Margin ----
    ws = wb.create_sheet('Profit & Margin')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Gross profit & margin, all history loaded so far (pre-tax on both revenue and cost - GST excluded as a pass-through). Qty in strips.'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    df_p = profit.rename(columns={'Qty_Sold_Strips': 'Qty Sold (strips)', 'Pretax_Revenue': 'Revenue (pre-tax)',
                                   'Cost_Per_Unit': 'Cost per Unit (pre-tax)', 'COGS': 'COGS (pre-tax)',
                                   'Gross_Profit': 'Gross Profit', 'Margin_Pct': 'Margin %'})
    df_p = df_p[['Product', 'Qty Sold (strips)', 'Revenue', 'Revenue (pre-tax)', 'Cost per Unit (pre-tax)',
                 'COGS (pre-tax)', 'Gross Profit', 'Margin %']]
    last_p = write_df(ws, df_p, start_row=2,
                       money_cols=['Revenue', 'Revenue (pre-tax)', 'Cost per Unit (pre-tax)', 'COGS (pre-tax)',
                                   'Gross Profit', 'Margin %'])

    r3 = last_p + 3
    ws.cell(row=r3, column=1,
            value='Cost unknown - sold with no purchase record in loaded history (from stock bought before this data started); excluded above').font = Font(name=FONT, bold=True, size=11)
    df_u = profit_unknown.rename(columns={'Qty_Sold_Strips': 'Qty Sold (strips)'})[['Product', 'Qty Sold (strips)', 'Revenue']]
    write_df(ws, df_u, start_row=r3 + 1, money_cols=['Revenue'])
    autosize(ws, [38, 12, 14, 16, 20, 16, 14, 10])
    ws.freeze_panes = 'A3'

    # ---- PTR Higher Than MRP ----
    ws = wb.create_sheet('PTR Higher Than MRP')
    ws.sheet_view.showGridLines = False
    if len(ptr_high) > 0:
        df3 = ptr_high[['Source_Month', 'Date', 'Inv.No', 'Supplier', 'Product', 'MRP', 'Sale Rate', 'Qty', 'Item Total', 'Excess']]
        df3 = df3.rename(columns={'Sale Rate': 'PTR', 'Source_Month': 'Month'})
    else:
        df3 = pd.DataFrame(columns=['Month', 'Date', 'Inv.No', 'Supplier', 'Product', 'MRP', 'PTR', 'Qty', 'Item Total', 'Excess'])
    write_df(ws, df3, money_cols=['MRP', 'PTR', 'Item Total', 'Excess'], qty_cols=['Qty'])
    autosize(ws, [12, 12, 14, 26, 32, 10, 10, 8, 12, 12])
    ws.freeze_panes = 'A2'

    # ---- MRP Issues ----
    ws = wb.create_sheet('MRP Issues')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f'MRP missing/zero this month ({latest_month})'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    if len(mrp_missing) > 0:
        df4a = mrp_missing[['Date', 'Inv.No', 'Supplier', 'Product', 'MRP', 'Sale Rate', 'Qty', 'Item Total']].rename(columns={'Sale Rate': 'PTR'})
    else:
        df4a = pd.DataFrame(columns=['Date', 'Inv.No', 'Supplier', 'Product', 'MRP', 'PTR', 'Qty', 'Item Total'])
    last = write_df(ws, df4a, start_row=2, money_cols=['MRP', 'PTR', 'Item Total'], qty_cols=['Qty'])

    r2 = last + 3
    ws.cell(row=r2, column=1, value='Same product, MRP varies 3x+ across all purchase history').font = Font(name=FONT, bold=True, size=11)
    df4b = variance.rename(columns={'min': 'Lowest MRP', 'max': 'Highest MRP', 'count': 'Purchase Lines', 'ratio': 'Spread (x)'})
    df4b['Spread (x)'] = df4b['Spread (x)'].round(1)
    write_df(ws, df4b, start_row=r2 + 1, money_cols=['Lowest MRP', 'Highest MRP'], qty_cols=['Purchase Lines'])
    autosize(ws, [14, 14, 26, 32, 12, 12, 8, 12])
    ws.freeze_panes = 'A3'

    # ---- Scheme Shortfall ----
    ws = wb.create_sheet('Scheme Shortfall')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f'Products where the usual free-qty scheme (e.g. 10+1, 10+2) was reduced or missing in {latest_month}'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    if len(scheme_missed) > 0:
        df5 = scheme_missed[['Date', 'Inv.No', 'Supplier', 'Product', 'Qty', 'Free Qty', 'FreeRatio',
                              'Typical_Ratio', 'Best_Ratio', 'Shortfall_Qty']].copy()
        df5['Free Qty'] = df5['Free Qty'].fillna(0)
        df5['FreeRatio'] = (df5['FreeRatio'].fillna(0) * 100).round(1)
        df5['Typical_Ratio'] = (df5['Typical_Ratio'] * 100).round(1)
        df5['Best_Ratio'] = (df5['Best_Ratio'] * 100).round(1)
        df5 = df5.rename(columns={'Qty': 'Purchased Qty', 'Free Qty': 'Free Qty Received',
                                   'FreeRatio': 'Free % Received', 'Typical_Ratio': 'Usual Free % (history)',
                                   'Best_Ratio': 'Best Free % Ever Seen', 'Shortfall_Qty': 'Shortfall (units)'})
    else:
        df5 = pd.DataFrame(columns=['Date', 'Inv.No', 'Supplier', 'Product', 'Purchased Qty', 'Free Qty Received',
                                     'Free % Received', 'Usual Free % (history)', 'Best Free % Ever Seen', 'Shortfall (units)'])
    write_df(ws, df5, start_row=2, qty_cols=['Purchased Qty', 'Free Qty Received', 'Shortfall (units)'],
             money_cols=['Free % Received', 'Usual Free % (history)', 'Best Free % Ever Seen'])
    autosize(ws, [12, 14, 26, 32, 12, 12, 12, 16, 16, 14])
    ws.freeze_panes = 'A3'

    # ---- Discount Shortfall ----
    ws = wb.create_sheet('Discount Shortfall')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f'Products getting a noticeably lower purchase discount % than their own history in {latest_month}'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    if len(disc_missed) > 0:
        df6 = disc_missed[['Date', 'Inv.No', 'Supplier', 'Product', 'Qty', 'Disc Percentage', 'Typical_Disc',
                            'Best_Disc', 'Disc_Gap_pct_pts', 'Value_Lost_Approx']].copy()
        df6 = df6.rename(columns={'Qty': 'Purchased Qty', 'Disc Percentage': 'Disc % Received',
                                   'Typical_Disc': 'Usual Disc % (history)', 'Best_Disc': 'Best Disc % Ever Seen',
                                   'Disc_Gap_pct_pts': 'Gap (pct pts)', 'Value_Lost_Approx': 'Approx Value Lost'})
    else:
        df6 = pd.DataFrame(columns=['Date', 'Inv.No', 'Supplier', 'Product', 'Purchased Qty', 'Disc % Received',
                                     'Usual Disc % (history)', 'Best Disc % Ever Seen', 'Gap (pct pts)', 'Approx Value Lost'])
    write_df(ws, df6, start_row=2, qty_cols=['Purchased Qty'],
             money_cols=['Disc % Received', 'Usual Disc % (history)', 'Best Disc % Ever Seen', 'Gap (pct pts)', 'Approx Value Lost'])
    autosize(ws, [12, 14, 26, 32, 12, 14, 16, 16, 12, 14])
    ws.freeze_panes = 'A3'

    # ---- Distributor Summary ----
    ws = wb.create_sheet('Distributor Summary')
    ws.sheet_view.showGridLines = False
    ws['A1'] = ('Potential profit embedded in what each distributor supplied, all history - if every unit received '
                '(incl. free-scheme goods) sold at MRP, vs what we paid. Buying/negotiation KPI, NOT realized profit.')
    ws['A1'].font = Font(name=FONT, italic=True, size=9, color='555555')
    df_ds = dist_summary.rename(columns={
        'Total_Invoice_Value': 'Total Invoice Value', 'Max_Sell_Pretax': 'Max Sell Value (pre-tax)',
        'Net_Cost_Pretax': 'Net Cost (pre-tax)', 'Embedded_Profit': 'Embedded Profit',
        'Margin_Pct': 'Margin %', 'Months_Active': 'Months Active'})
    df_ds = df_ds[['Supplier', 'Invoices', 'Lines', 'Total Invoice Value', 'Max Sell Value (pre-tax)',
                    'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %', 'Months Active']]
    write_df(ws, df_ds, start_row=2,
             money_cols=['Total Invoice Value', 'Max Sell Value (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
             qty_cols=['Invoices', 'Lines', 'Months Active'])
    autosize(ws, [34, 11, 10, 18, 20, 18, 16, 10, 12])
    ws.freeze_panes = 'A3'

    # ---- Distributor x Month ----
    ws = wb.create_sheet('Distributor x Month')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Supplier x Month detail'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    df_dm = dist_month.rename(columns={'Source_Month': 'Month', 'Invoice_Value': 'Invoice Value',
                                        'Max_Sell_Pretax': 'Max Sell (pre-tax)', 'Net_Cost_Pretax': 'Net Cost (pre-tax)',
                                        'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
    df_dm = df_dm[['Supplier', 'Month', 'Invoice Value', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
    df_dm['Supplier'] = df_dm['Supplier'].astype(str)
    last_dm = write_df(ws, df_dm, start_row=2,
                        money_cols=['Invoice Value', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'])

    r_piv = last_dm + 3
    ws.cell(row=r_piv, column=1, value='Pivot: Embedded Profit by Supplier x Month (suppliers ordered by all-history total)').font = \
        Font(name=FONT, bold=True, size=11)
    dist_pivot = dist_month.pivot(index='Supplier', columns='Source_Month', values='Embedded_Profit').fillna(0)
    dist_pivot = dist_pivot.reindex(dist_supplier_order)
    dist_pivot['Total'] = dist_pivot.sum(axis=1)
    dist_pivot = dist_pivot.reset_index()
    dist_pivot['Supplier'] = dist_pivot['Supplier'].astype(str)
    write_df(ws, dist_pivot, start_row=r_piv + 1, money_cols=list(dist_pivot.columns[1:]))
    autosize(ws, [34, 16, 16, 16, 16, 16])
    ws.freeze_panes = 'A3'

    # ---- Distributor Per Invoice ----
    ws = wb.create_sheet('Distributor Per Invoice')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Potential profit per individual purchase invoice, biggest first'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    df_pi = dist_per_invoice.rename(columns={
        'Source_Month': 'Month', 'Invoice_Value': 'Invoice Value', 'Max_Sell_Pretax': 'Max Sell (pre-tax)',
        'Net_Cost_Pretax': 'Net Cost (pre-tax)', 'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
    df_pi = df_pi[['Date', 'Inv.No', 'Supplier', 'Month', 'Lines', 'Invoice Value',
                    'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
    write_df(ws, df_pi, start_row=2,
             money_cols=['Invoice Value', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
             qty_cols=['Lines'])
    autosize(ws, [12, 16, 30, 10, 8, 16, 16, 16, 16, 10])
    ws.freeze_panes = 'A3'

    # ---- Top Products by Distributor ----
    ws = wb.create_sheet('Top Products by Distributor')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f'Top {TOP_N_PRODUCTS_PER_DISTRIBUTOR} products by embedded profit, per distributor (suppliers ordered by all-history total)'
    ws['A1'].font = Font(name=FONT, bold=True, size=11)
    df_tp = dist_top_products.rename(columns={
        'Units_Received': 'Units Received', 'Max_Sell_Pretax': 'Max Sell (pre-tax)',
        'Net_Cost_Pretax': 'Net Cost (pre-tax)', 'Embedded_Profit': 'Embedded Profit', 'Margin_Pct': 'Margin %'})
    df_tp = df_tp[['Supplier', 'Product', 'Units Received', 'Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %']]
    write_df(ws, df_tp, start_row=2,
             money_cols=['Max Sell (pre-tax)', 'Net Cost (pre-tax)', 'Embedded Profit', 'Margin %'],
             qty_cols=['Units Received'])
    autosize(ws, [30, 32, 14, 16, 16, 16, 10])
    ws.freeze_panes = 'A3'

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, latest_month, target_month


def main():
    sales, purch = ingest()
    if sales.empty or purch.empty:
        print('Need at least one Sale file and one Purchase file in data\\raw to run.')
        return
    out_dir = REPORTS_DIR / sorted(sales['Source_Month'].unique())[-1]
    out_path = out_dir / 'Monthly Reckoner Report.xlsx'
    path, latest_month, target_month = build_report(sales, purch, out_path)
    print()
    print(f'Report saved: {path}')
    print(f'History months: {sorted(sales["Source_Month"].unique())}')
    print(f'Latest month analyzed for issues: {latest_month}')
    print(f'Forecasting: {target_month}')


if __name__ == '__main__':
    main()
